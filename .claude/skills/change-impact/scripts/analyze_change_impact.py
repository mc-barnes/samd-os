#!/usr/bin/env python3
"""
Change Impact Analyzer for SaMD Regulatory Compliance.

Generates XLSX change impact reports with traceability to design controls
and risk analysis, re-verification scope, and regulatory pathway assessment.

Usage:
    python scripts/analyze_change_impact.py --example spo2
    python scripts/analyze_change_impact.py --example spo2 --design-controls dc.xlsx --risk-analysis ra.xlsx
"""

import argparse
import os
import sys
from datetime import date
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


# --- Style Constants ---
HEADER_FILL = PatternFill(start_color="CC6600", end_color="CC6600", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
MINOR_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
MAJOR_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
CRITICAL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")


def get_spo2_example():
    """Return hardcoded SpO2 eval pipeline change impact example data."""
    change_request = {
        "cr_id": "CR-001",
        "description": (
            "Introduce new emergency classification tier for SpO2 readings "
            "below 80%. Currently system only classifies normal/borderline/urgent. "
            "Emergency tier triggers immediate page."
        ),
        "files_changed": [
            "src/rules/classifier.py",
            "src/pipeline/orchestrator.py",
            "src/interop/hl7_messages.py",
            "app/components/trace_viewer.py",
        ],
        "rationale": (
            "Clinical safety review identified gap in handling critical "
            "desaturation events below 80% SpO2"
        ),
        "date": date.today().isoformat(),
        "requestor": "Clinical Safety Team",
    }

    impact_trace = [
        {
            "changed_file": "src/rules/classifier.py",
            "affected_di_ids": "DI-001 (classification), DI-002 (desat detection)",
            "affected_haz_ids": "HAZ-001 (misclassify urgent)",
            "affected_ver_ids": "VER-001, VER-002",
            "notes": "Core classification logic modified to add emergency tier",
        },
        {
            "changed_file": "src/pipeline/orchestrator.py",
            "affected_di_ids": "DI-001 (classification)",
            "affected_haz_ids": "HAZ-001 (misclassify urgent)",
            "affected_ver_ids": "VER-001",
            "notes": "Pipeline routing updated for new tier",
        },
        {
            "changed_file": "src/interop/hl7_messages.py",
            "affected_di_ids": "DI-003 (handoff reports)",
            "affected_haz_ids": "HAZ-005 (handoff context)",
            "affected_ver_ids": "VER-003",
            "notes": "HL7 message templates updated for emergency tier",
        },
        {
            "changed_file": "app/components/trace_viewer.py",
            "affected_di_ids": "DI-003 (handoff reports)",
            "affected_haz_ids": "—",
            "affected_ver_ids": "VER-003",
            "notes": "UI display updated (no direct safety impact)",
        },
    ]

    reverification_scope = [
        {
            "ver_id": "VER-001",
            "test_name": "Urgency classification test suite",
            "rationale": "Classification logic changed",
            "effort_estimate": "4 hours",
            "priority": "P1",
        },
        {
            "ver_id": "VER-002",
            "test_name": "Desaturation detection accuracy",
            "rationale": "New threshold added",
            "effort_estimate": "2 hours",
            "priority": "P1",
        },
        {
            "ver_id": "VER-003",
            "test_name": "Handoff report completeness",
            "rationale": "New tier in report output",
            "effort_estimate": "1 hour",
            "priority": "P2",
        },
    ]

    regulatory_assessment = {
        "classification": "Critical",
        "pathway": "Regulatory review required — likely new 510(k) or Letter to File with extensive V&V",
        "rationale": (
            "Adds new safety-critical classification tier (emergency) that triggers "
            "immediate clinical action (page). New clinical decision point affecting "
            "patient safety. Per FDA guidance: new risk control implementation in "
            "safety-critical pathway. If device has cleared predicate, formal 510(k) "
            "equivalence assessment required to determine if new submission needed."
        ),
        "new_510k": "Requires assessment — see Regulatory_Assessment",
        "risk_analysis_update": (
            "Yes — add HAZ for emergency tier misclassification"
        ),
        "cumulative_impact_notes": (
            "First major change since initial release. No cumulative threshold concern."
        ),
        # 510(k) Equivalence Assessment fields
        "predicate_device": "N/A — portfolio demonstration project",
        "predicate_intended_use": "N/A",
        "performance_claims_affected": "N/A — no cleared predicate",
        "se_conclusion": "N/A",
        "evidence_required": (
            "Full system V&V including emergency tier classification accuracy, "
            "alarm latency testing, clinical workflow validation"
        ),
    }

    return change_request, impact_trace, reverification_scope, regulatory_assessment


def try_read_design_controls(path):
    """Attempt to read DI IDs from a design controls XLSX file."""
    try:
        wb = load_workbook(path, data_only=True)
        if "Design_Inputs" in wb.sheetnames:
            ws = wb["Design_Inputs"]
            di_ids = []
            for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
                if row[0]:
                    di_id = str(row[0])
                    desc = str(row[1]) if row[1] else ""
                    di_ids.append(f"{di_id} ({desc})" if desc else di_id)
            wb.close()
            return di_ids
        wb.close()
    except Exception as e:
        print(f"  Warning: Could not read design controls from {path}: {e}")
        print("  Falling back to template IDs.")
    return None


def try_read_risk_analysis(path):
    """Attempt to read HAZ IDs from a risk analysis XLSX file."""
    try:
        wb = load_workbook(path, data_only=True)
        if "Hazard_Identification" in wb.sheetnames:
            ws = wb["Hazard_Identification"]
            haz_ids = []
            for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
                if row[0]:
                    haz_id = str(row[0])
                    desc = str(row[1]) if row[1] else ""
                    haz_ids.append(f"{haz_id} ({desc})" if desc else haz_id)
            wb.close()
            return haz_ids
        wb.close()
    except Exception as e:
        print(f"  Warning: Could not read risk analysis from {path}: {e}")
        print("  Falling back to template IDs.")
    return None


def style_header_row(ws, headers):
    """Apply header styling to the first row of a worksheet."""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP_ALIGNMENT


def auto_adjust_column_widths(ws):
    """Auto-adjust column widths based on content."""
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                # Account for line breaks in wrapped text
                lines = str(cell.value).split("\n")
                longest_line = max(len(line) for line in lines)
                max_length = max(max_length, longest_line)
        adjusted_width = min(max(max_length + 2, 12), 60)
        ws.column_dimensions[col_letter].width = adjusted_width


def get_classification_fill(classification):
    """Return the appropriate fill color for a classification level."""
    level = classification.lower()
    if level == "minor":
        return MINOR_FILL
    elif level == "major":
        return MAJOR_FILL
    elif level == "critical":
        return CRITICAL_FILL
    return None


def build_document_control_sheet(wb, device_name):
    """Create and populate the Document_Control sheet as the first sheet."""
    ws = wb.active
    ws.title = "Document_Control"

    # Row 1: Merged header
    ws.merge_cells("A1:D1")
    cell = ws.cell(row=1, column=1, value="Document Control")
    cell.fill = HEADER_FILL
    cell.font = Font(bold=True, color="FFFFFF", size=12)
    cell.alignment = Alignment(horizontal="center", vertical="center")

    # Row 3-6: Document metadata
    ws.cell(row=3, column=1, value="Document Title").font = Font(bold=True)
    ws.cell(row=3, column=2, value=f"Change Impact Analysis — {device_name}")

    ws.cell(row=4, column=1, value="Document Version").font = Font(bold=True)
    ws.cell(row=4, column=2, value="1.0")

    ws.cell(row=5, column=1, value="Device/Software Version").font = Font(bold=True)
    if "spo2" in device_name.lower():
        ws.cell(row=5, column=2, value="SpO2 AI Eval Pipeline v2.1")
    else:
        ws.cell(row=5, column=2, value=device_name)

    ws.cell(row=6, column=1, value="Date Generated").font = Font(bold=True)
    ws.cell(row=6, column=2, value=date.today().isoformat())

    # Rows 8-11: QMS integration fields
    qms_fields = [
        ("QMS Document Number", ""),
        ("SOP Reference", ""),
        ("Related Design Controls File", ""),
        ("Related Risk Management File", ""),
    ]
    for idx, (label, value) in enumerate(qms_fields):
        row = 8 + idx
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value).alignment = WRAP_ALIGNMENT

    # Row 13: Approval Signatures header
    ws.merge_cells("A13:D13")
    cell = ws.cell(row=13, column=1, value="Approval Signatures")
    cell.font = Font(bold=True, size=11)

    # Row 14: Signature table headers
    sig_headers = ["Role", "Name", "Date", "Signature"]
    for col_idx, header in enumerate(sig_headers, 1):
        cell = ws.cell(row=14, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP_ALIGNMENT

    # Rows 15-18: Signature rows
    roles = [
        "Prepared By",
        "Reviewed By (QA)",
        "Approved By (Management)",
        "Approved By (Regulatory)",
    ]
    for row_idx, role in enumerate(roles, 15):
        ws.cell(row=row_idx, column=1, value=role).alignment = WRAP_ALIGNMENT
        ws.cell(row=row_idx, column=2, value="").alignment = WRAP_ALIGNMENT
        ws.cell(row=row_idx, column=3, value="").alignment = WRAP_ALIGNMENT
        ws.cell(row=row_idx, column=4, value="").alignment = WRAP_ALIGNMENT

    auto_adjust_column_widths(ws)
    return ws


def build_change_request_sheet(wb, change_request):
    """Create and populate the Change_Request sheet."""
    ws = wb.create_sheet("Change_Request")

    headers = ["CR ID", "Description", "Files Changed", "Rationale", "Date", "Requestor"]
    style_header_row(ws, headers)

    row_data = [
        change_request["cr_id"],
        change_request["description"],
        ", ".join(change_request["files_changed"]),
        change_request["rationale"],
        change_request["date"],
        change_request["requestor"],
    ]
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=2, column=col_idx, value=value)
        cell.alignment = WRAP_ALIGNMENT

    # Summary formula: total change requests
    ws.cell(row=4, column=1, value="Total CRs:")
    ws.cell(row=4, column=2, value="=COUNTA(A2:A100)")

    auto_adjust_column_widths(ws)
    return ws


def build_impact_trace_sheet(wb, impact_trace):
    """Create and populate the Impact_Trace sheet."""
    ws = wb.create_sheet("Impact_Trace")

    headers = [
        "Changed File",
        "Affected DI IDs",
        "Affected HAZ IDs",
        "Affected VER IDs",
        "Notes",
    ]
    style_header_row(ws, headers)

    for row_idx, trace in enumerate(impact_trace, 2):
        ws.cell(row=row_idx, column=1, value=trace["changed_file"]).alignment = WRAP_ALIGNMENT
        ws.cell(row=row_idx, column=2, value=trace["affected_di_ids"]).alignment = WRAP_ALIGNMENT
        ws.cell(row=row_idx, column=3, value=trace["affected_haz_ids"]).alignment = WRAP_ALIGNMENT
        ws.cell(row=row_idx, column=4, value=trace["affected_ver_ids"]).alignment = WRAP_ALIGNMENT
        ws.cell(row=row_idx, column=5, value=trace["notes"]).alignment = WRAP_ALIGNMENT

    # Summary formula: total affected items
    summary_row = len(impact_trace) + 3
    ws.cell(row=summary_row, column=1, value="Total Affected Files:")
    ws.cell(row=summary_row, column=2, value='=COUNTA(A2:A100)')
    ws.cell(row=summary_row + 1, column=1, value="Affected Items (non-empty):")
    ws.cell(row=summary_row + 1, column=2, value='=SUMPRODUCT((D2:D10<>"")*1)')

    auto_adjust_column_widths(ws)
    return ws


def build_reverification_sheet(wb, reverification_scope):
    """Create and populate the Reverification_Scope sheet."""
    ws = wb.create_sheet("Reverification_Scope")

    headers = ["VER ID", "Test Name", "Rationale", "Effort Estimate", "Priority"]
    style_header_row(ws, headers)

    for row_idx, ver in enumerate(reverification_scope, 2):
        ws.cell(row=row_idx, column=1, value=ver["ver_id"]).alignment = WRAP_ALIGNMENT
        ws.cell(row=row_idx, column=2, value=ver["test_name"]).alignment = WRAP_ALIGNMENT
        ws.cell(row=row_idx, column=3, value=ver["rationale"]).alignment = WRAP_ALIGNMENT
        ws.cell(row=row_idx, column=4, value=ver["effort_estimate"]).alignment = WRAP_ALIGNMENT
        ws.cell(row=row_idx, column=5, value=ver["priority"]).alignment = WRAP_ALIGNMENT

    # Summary
    summary_row = len(reverification_scope) + 3
    ws.cell(row=summary_row, column=1, value="Total Verifications:")
    ws.cell(row=summary_row, column=2, value="=COUNTA(A2:A100)")

    auto_adjust_column_widths(ws)
    return ws


def build_regulatory_assessment_sheet(wb, regulatory_assessment):
    """Create and populate the Regulatory_Assessment sheet."""
    ws = wb.create_sheet("Regulatory_Assessment")

    headers = ["Field", "Value"]
    style_header_row(ws, headers)

    fields = [
        ("Classification", regulatory_assessment["classification"]),
        ("Pathway", regulatory_assessment["pathway"]),
        ("Rationale", regulatory_assessment["rationale"]),
        ("New 510(k) Required", regulatory_assessment["new_510k"]),
        ("Risk Analysis Update Required", regulatory_assessment["risk_analysis_update"]),
        ("Cumulative Impact Notes", regulatory_assessment["cumulative_impact_notes"]),
    ]

    for row_idx, (field, value) in enumerate(fields, 2):
        ws.cell(row=row_idx, column=1, value=field).alignment = WRAP_ALIGNMENT
        cell = ws.cell(row=row_idx, column=2, value=value)
        cell.alignment = WRAP_ALIGNMENT

        # Apply conditional color to classification cell
        if field == "Classification":
            fill = get_classification_fill(value)
            if fill:
                cell.fill = fill
                cell.font = Font(bold=True, size=11)

    # 510(k) Equivalence Assessment sub-section
    equiv_start = len(fields) + 4  # blank row, then sub-header
    ws.merge_cells(f"A{equiv_start}:B{equiv_start}")
    cell = ws.cell(row=equiv_start, column=1, value="510(k) Equivalence Assessment")
    cell.font = Font(bold=True, size=11)

    equiv_fields = [
        (
            "Predicate Device",
            "510(k) K-number or 'First of kind'",
            regulatory_assessment.get("predicate_device", ""),
        ),
        (
            "Predicate Intended Use",
            "Same / Different / Expanded",
            regulatory_assessment.get("predicate_intended_use", ""),
        ),
        (
            "Performance Claims Affected",
            "Yes / No — does change violate cleared performance claims?",
            regulatory_assessment.get("performance_claims_affected", ""),
        ),
        (
            "Substantial Equivalence Conclusion",
            "Equivalent / Not Equivalent / Assessment Required",
            regulatory_assessment.get("se_conclusion", ""),
        ),
        (
            "Evidence Required",
            "List V&V evidence needed for equivalence determination",
            regulatory_assessment.get("evidence_required", ""),
        ),
    ]

    for i, (field, description, value) in enumerate(equiv_fields):
        row = equiv_start + 1 + i
        ws.cell(row=row, column=1, value=field).alignment = WRAP_ALIGNMENT
        cell = ws.cell(row=row, column=2, value=value if value else description)
        cell.alignment = WRAP_ALIGNMENT

    auto_adjust_column_widths(ws)
    return ws


def main():
    parser = argparse.ArgumentParser(
        description="Change Impact Analyzer for SaMD Regulatory Compliance"
    )
    parser.add_argument(
        "--device-name",
        type=str,
        default=None,
        help="Device/product name for output file naming",
    )
    parser.add_argument(
        "--example",
        type=str,
        choices=["spo2"],
        help="Use a built-in example (spo2 = SpO2 AI Eval Pipeline)",
    )
    parser.add_argument(
        "--design-controls",
        type=str,
        default=None,
        help="Path to design controls XLSX (optional, from design-controls skill)",
    )
    parser.add_argument(
        "--risk-analysis",
        type=str,
        default=None,
        help="Path to risk analysis XLSX (optional, from risk-management skill)",
    )
    args = parser.parse_args()

    if not args.example:
        parser.error("--example is required (currently only 'spo2' is supported)")

    # Determine device name
    if args.device_name:
        device_name = args.device_name
    elif args.example == "spo2":
        device_name = "spo2-ai-eval-pipeline"
    else:
        device_name = "unknown-device"

    print(f"Change Impact Analyzer")
    print(f"{'=' * 50}")
    print(f"Device: {device_name}")
    print(f"Example: {args.example}")
    print()

    # Load example data
    if args.example == "spo2":
        change_request, impact_trace, reverification_scope, regulatory_assessment = (
            get_spo2_example()
        )

    # Optionally enrich from upstream skill outputs
    if args.design_controls:
        print(f"Reading design controls from: {args.design_controls}")
        di_ids = try_read_design_controls(args.design_controls)
        if di_ids:
            print(f"  Found {len(di_ids)} Design Input IDs")
        else:
            print("  Using template DI IDs")
        print()

    if args.risk_analysis:
        print(f"Reading risk analysis from: {args.risk_analysis}")
        haz_ids = try_read_risk_analysis(args.risk_analysis)
        if haz_ids:
            print(f"  Found {len(haz_ids)} Hazard IDs")
        else:
            print("  Using template HAZ IDs")
        print()

    # Build workbook
    print("Building XLSX report...")
    wb = Workbook()

    build_document_control_sheet(wb, device_name)
    build_change_request_sheet(wb, change_request)
    build_impact_trace_sheet(wb, impact_trace)
    build_reverification_sheet(wb, reverification_scope)
    build_regulatory_assessment_sheet(wb, regulatory_assessment)

    # Save output
    output_dir = Path(__file__).parent.parent / "output"
    output_dir.mkdir(exist_ok=True)
    output_path = output_dir / f"change-impact-{device_name}.xlsx"
    wb.save(str(output_path))
    print(f"  Saved: {output_path}")
    print()

    # Summary
    print("Report Summary:")
    print(f"  Change Requests: 1")
    print(f"  Files Impacted: {len(impact_trace)}")
    print(f"  Verifications to Re-execute: {len(reverification_scope)}")
    print(f"  Classification: {regulatory_assessment['classification']}")
    print(f"  Pathway: {regulatory_assessment['pathway']}")
    print(f"  New 510(k): {regulatory_assessment['new_510k']}")
    print()
    print("Done.")


if __name__ == "__main__":
    main()
