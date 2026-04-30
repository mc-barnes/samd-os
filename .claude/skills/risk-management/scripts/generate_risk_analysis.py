#!/usr/bin/env python3
"""
ISO 14971 Risk Management XLSX Generator

Generates an 8-sheet risk analysis workbook:
  1. Document_Control       (metadata & approval signatures)
  2. Hazard_Identification
  3. Risk_Estimation
  4. Risk_Evaluation
  5. Risk_Controls
  6. Residual_Risk          (with ALARP rationale enforcement)
  7. Overall_Residual_Risk  (ISO 14971 Clause 7)
  8. FMEA

Uses real Excel formulas for risk level lookups and RPN calculations.
Standalone -- requires only openpyxl.

Usage:
    python scripts/generate_risk_analysis.py --example spo2
    python scripts/generate_risk_analysis.py --device-name "My Device"
"""

import argparse
import os
import re
from datetime import date
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CELL_ALIGNMENT = Alignment(vertical="top", wrap_text=True)

# Conditional-formatting fills
RED_FILL = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
GREEN_FILL = PatternFill(start_color="66BB6A", end_color="66BB6A", fill_type="solid")

# 5x5 risk acceptability matrix  [severity_index][probability_index]
# Row 0 = S1 (Negligible) ... Row 4 = S5 (Catastrophic)
# Col 0 = P1 (Incredible)  ... Col 4 = P5 (Frequent)
RISK_MATRIX: list[list[str]] = [
    # P1       P2       P3       P4       P5
    ["AC",    "AC",    "AC",    "AC",    "A"],      # S1
    ["AC",    "AC",    "A",     "A",     "U"],      # S2
    ["AC",    "A",     "A",     "U",     "U"],      # S3
    ["A",     "A",     "U",     "U",     "U"],      # S4
    ["A",     "U",     "U",     "U",     "U"],      # S5
]

ACCEPTABILITY_MAP = {
    "AC": "Acceptable",
    "A": "ALARP",
    "U": "Unacceptable",
}


# ---------------------------------------------------------------------------
# SpO2 example data
# ---------------------------------------------------------------------------
def _spo2_hazards() -> list[dict[str, Any]]:
    """Return the 5 SpO2 AI example hazards."""
    return [
        {
            "id": "HAZ-001",
            "hazard": "Misclassify urgent desaturation as normal",
            "situation": "Algorithm classifies a true urgent desaturation event as normal SpO2, and no alarm is triggered",
            "harm": "Delayed clinical intervention, potential brain injury",
            "sequence": "Sensor captures low SpO2 -> AI model misclassifies severity -> no alarm fired -> clinician unaware -> delayed O2 therapy -> hypoxic brain injury",
            "severity": 4, "probability": 3,
            "rc_id": "RC-001",
            "control": "Dual-threshold validation with GA-adjusted limits",
            "control_type": "Inherent safety by design",
            "verification": "Unit tests with edge-case SpO2 waveforms; clinical validation against expert labels (n>=200)",
            "post_severity": 4, "post_probability": 1,
            "rationale": "Residual risk is ALARP: dual-threshold with GA adjustment reduces probability to Incredible; further reduction requires fundamentally different sensor technology which is impracticable",
            "fm": "Classification algorithm outputs 'normal' for SpO2 < 85%",
            "effect": "No alarm triggered for urgent desaturation",
            "cause": "Model undertrained on preterm-specific waveforms",
            "fmea_s": 8, "fmea_o": 4, "fmea_d": 3,
        },
        {
            "id": "HAZ-002",
            "hazard": "False positive artifact flag on real desaturation",
            "situation": "Real desaturation event is flagged as motion artifact and discarded from analysis",
            "harm": "Real desat discarded, delayed clinical response",
            "sequence": "True desat occurs -> artifact filter misidentifies as noise -> event dropped -> clinician not notified -> delayed response",
            "severity": 3, "probability": 4,
            "rc_id": "RC-002",
            "control": "Confidence scoring with expert queue fallback",
            "control_type": "Protective measure",
            "verification": "ROC analysis on artifact vs. real desat test set; expert review queue audit (monthly)",
            "post_severity": 3, "post_probability": 2,
            "rationale": "Residual risk is ALARP: confidence scoring routes uncertain cases to expert review; residual probability is Improbable",
            "fm": "Artifact filter rejects waveform segments with SpO2 < 88%",
            "effect": "Desaturation event not recorded or alarmed",
            "cause": "Artifact filter threshold too aggressive for low-signal conditions",
            "fmea_s": 6, "fmea_o": 5, "fmea_d": 4,
        },
        {
            "id": "HAZ-003",
            "hazard": "Sensor disconnect undetected",
            "situation": "Pulse oximetry sensor becomes disconnected and system does not alert nursing staff",
            "harm": "No monitoring during disconnect period",
            "sequence": "Sensor dislodged -> signal loss -> system does not detect disconnect -> monitoring gap -> event missed",
            "severity": 3, "probability": 3,
            "rc_id": "RC-003",
            "control": "Signal continuity check every 5 seconds",
            "control_type": "Protective measure",
            "verification": "Simulated disconnect test (n=50); verify alarm triggers within 10s of signal loss",
            "post_severity": 3, "post_probability": 1,
            "rationale": "Residual risk is Acceptable: continuous signal check reduces probability to Incredible",
            "fm": "System does not generate disconnect alarm within 30s",
            "effect": "Monitoring gap during sensor disconnect",
            "cause": "Signal loss detection logic has timeout > clinical threshold",
            "fmea_s": 6, "fmea_o": 3, "fmea_d": 2,
        },
        {
            "id": "HAZ-004",
            "hazard": "Wrong GA threshold applied",
            "situation": "Gestational age category is incorrectly assigned, causing wrong SpO2 thresholds to be used for classification",
            "harm": "Inappropriate urgency classification",
            "sequence": "GA entered incorrectly or defaulted -> wrong threshold table selected -> classification uses wrong limits -> over/under-triage",
            "severity": 3, "probability": 2,
            "rc_id": "RC-004",
            "control": "GA category validation at admission",
            "control_type": "Inherent safety by design",
            "verification": "Input validation unit tests; GA boundary-value testing; admission workflow audit (quarterly)",
            "post_severity": 3, "post_probability": 1,
            "rationale": "Residual risk is Acceptable: validation at admission with boundary checks reduces probability to Incredible",
            "fm": "GA lookup returns incorrect category for edge-case gestational ages",
            "effect": "SpO2 thresholds do not match patient's clinical profile",
            "cause": "Off-by-one error in GA week boundary logic",
            "fmea_s": 6, "fmea_o": 3, "fmea_d": 2,
        },
        {
            "id": "HAZ-005",
            "hazard": "Handoff report omits critical context",
            "situation": "AI-generated shift handoff report is missing key clinical context needed for care continuity",
            "harm": "Clinician makes suboptimal decision due to missing context",
            "sequence": "Handoff report generated -> critical event omitted -> receiving clinician unaware of history -> suboptimal care decision",
            "severity": 2, "probability": 3,
            "rc_id": "RC-005",
            "control": "Structured template with mandatory fields",
            "control_type": "Information for safety",
            "verification": "Template completeness validation; clinician usability testing (n>=10); monthly content audit",
            "post_severity": 2, "post_probability": 2,
            "rationale": "Residual risk is Acceptable: mandatory fields ensure critical data inclusion; probability reduced to Improbable",
            "fm": "Handoff template renders with empty mandatory fields",
            "effect": "Clinician receives incomplete handoff information",
            "cause": "Template logic does not enforce required field population",
            "fmea_s": 4, "fmea_o": 4, "fmea_d": 3,
        },
    ]


# ---------------------------------------------------------------------------
# Helper: style a header row
# ---------------------------------------------------------------------------
def _style_headers(ws: Worksheet, headers: list[str], col_widths: list[int]) -> None:
    """Write header row with dark-red background and set column widths."""
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"


def _style_cell(ws: Worksheet, row: int, col: int, value: Any) -> None:
    """Write a value and apply standard cell styling."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.alignment = CELL_ALIGNMENT
    cell.border = THIN_BORDER


# ---------------------------------------------------------------------------
# Risk level Excel formula
# ---------------------------------------------------------------------------
def _risk_level_formula(sev_cell: str, prob_cell: str) -> str:
    """
    Build a nested-IF Excel formula that looks up the 5x5 risk matrix.

    Given a severity cell (e.g. B2) containing 1-5 and a probability cell
    (e.g. C2) containing 1-5, returns "Acceptable", "ALARP", or
    "Unacceptable".
    """
    # Build the formula row by row. Each row checks severity, then
    # probability within that severity level.
    parts: list[str] = []
    for s_idx in range(5):
        s_val = s_idx + 1  # 1-based severity
        inner_parts: list[str] = []
        for p_idx in range(5):
            p_val = p_idx + 1
            result = ACCEPTABILITY_MAP[RISK_MATRIX[s_idx][p_idx]]
            inner_parts.append(
                f'IF({prob_cell}={p_val},"{result}"'
            )
        # Chain the probability IFs for this severity level
        # =IF(P=1,"AC",IF(P=2,"AC",IF(P=3,"AC",IF(P=4,"AC",IF(P=5,"A","")))))
        inner_formula = ""
        for i, ip in enumerate(reversed(inner_parts)):
            if i == 0:
                inner_formula = f'{ip},"ERROR"{")" * 1}'
            else:
                inner_formula = f"{ip},{inner_formula})"
        parts.append(f"IF({sev_cell}={s_val},{inner_formula}")

    # Chain the severity IFs
    formula = ""
    for i, part in enumerate(reversed(parts)):
        if i == 0:
            formula = f'{part},"ERROR"{")" * 1}'
        else:
            formula = f"{part},{formula})"

    return f"={formula}"


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------
def _build_document_control(ws: Worksheet, device_name: str) -> None:
    """Build the Document_Control sheet with metadata and signature block."""
    col_widths = [28, 45, 16, 16]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Row 1: Merged header
    ws.merge_cells("A1:D1")
    cell = ws.cell(row=1, column=1, value="Document Control")
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = HEADER_ALIGNMENT
    cell.border = THIN_BORDER
    for col in range(2, 5):
        c = ws.cell(row=1, column=col)
        c.fill = HEADER_FILL
        c.border = THIN_BORDER

    # Rows 3-6: Document metadata
    meta_rows = [
        ("Document Title", f"Risk Analysis — {device_name}"),
        ("Document Version", "1.0"),
        ("Device/Software Version", device_name),
        ("Date Generated", date.today().isoformat()),
    ]
    for row_idx, (label, value) in enumerate(meta_rows, start=3):
        label_cell = ws.cell(row=row_idx, column=1, value=label)
        label_cell.font = Font(bold=True, size=11)
        label_cell.alignment = CELL_ALIGNMENT
        label_cell.border = THIN_BORDER
        val_cell = ws.cell(row=row_idx, column=2, value=value)
        val_cell.alignment = CELL_ALIGNMENT
        val_cell.border = THIN_BORDER

    # Rows 8-10: QMS integration fields
    qms_fields = [
        ("QMS Document Number", ""),
        ("SOP Reference", ""),
        ("Related Design Controls File", ""),
    ]
    for idx, (label, value) in enumerate(qms_fields):
        row = 8 + idx
        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font = Font(bold=True, size=11)
        label_cell.alignment = CELL_ALIGNMENT
        label_cell.border = THIN_BORDER
        val_cell = ws.cell(row=row, column=2, value=value)
        val_cell.alignment = CELL_ALIGNMENT
        val_cell.border = THIN_BORDER

    # Row 12: Approval Signatures header (merged)
    ws.merge_cells("A12:D12")
    sig_header = ws.cell(row=12, column=1, value="Approval Signatures")
    sig_header.font = Font(bold=True, size=11)
    sig_header.alignment = CELL_ALIGNMENT
    sig_header.border = THIN_BORDER
    for col in range(2, 5):
        c = ws.cell(row=12, column=col)
        c.border = THIN_BORDER

    # Row 13: Column headers for signatures
    sig_headers = ["Role", "Name", "Date", "Signature"]
    for col_idx, header in enumerate(sig_headers, start=1):
        cell = ws.cell(row=13, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Rows 14-16: Signature rows
    sig_roles = ["Prepared By", "Reviewed By (QA)", "Approved By (Management)"]
    for row_idx, role in enumerate(sig_roles, start=14):
        _style_cell(ws, row_idx, 1, role)
        for col in range(2, 5):
            _style_cell(ws, row_idx, col, "")


def _build_hazard_identification(wb: Workbook, hazards: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Hazard_Identification")
    headers = ["Hazard ID", "Hazard", "Hazardous Situation", "Harm", "Sequence of Events"]
    widths = [12, 40, 45, 40, 55]
    _style_headers(ws, headers, widths)

    for i, h in enumerate(hazards, start=2):
        _style_cell(ws, i, 1, h["id"])
        _style_cell(ws, i, 2, h["hazard"])
        _style_cell(ws, i, 3, h["situation"])
        _style_cell(ws, i, 4, h["harm"])
        _style_cell(ws, i, 5, h["sequence"])


def _build_risk_estimation(wb: Workbook, hazards: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Risk_Estimation")
    headers = ["Hazard ID", "Severity (S1-S5)", "Probability (P1-P5)", "Risk Level"]
    widths = [12, 18, 20, 18]
    _style_headers(ws, headers, widths)

    # Data validation for severity and probability
    sev_dv = DataValidation(type="whole", operator="between", formula1="1", formula2="5")
    sev_dv.error = "Severity must be between 1 (Negligible) and 5 (Catastrophic)"
    sev_dv.errorTitle = "Invalid Severity"
    prob_dv = DataValidation(type="whole", operator="between", formula1="1", formula2="5")
    prob_dv.error = "Probability must be between 1 (Incredible) and 5 (Frequent)"
    prob_dv.errorTitle = "Invalid Probability"
    ws.add_data_validation(sev_dv)
    ws.add_data_validation(prob_dv)

    for i, h in enumerate(hazards, start=2):
        _style_cell(ws, i, 1, h["id"])
        _style_cell(ws, i, 2, h["severity"])
        _style_cell(ws, i, 3, h["probability"])
        # Risk Level formula
        formula = _risk_level_formula(f"B{i}", f"C{i}")
        cell = ws.cell(row=i, column=4, value=formula)
        cell.alignment = CELL_ALIGNMENT
        cell.border = THIN_BORDER

        sev_dv.add(ws.cell(row=i, column=2))
        prob_dv.add(ws.cell(row=i, column=3))

    # Conditional formatting on Risk Level (column D)
    risk_range = f"D2:D{len(hazards) + 1}"
    ws.conditional_formatting.add(
        risk_range,
        CellIsRule(operator="equal", formula=['"Unacceptable"'], fill=RED_FILL),
    )
    ws.conditional_formatting.add(
        risk_range,
        CellIsRule(operator="equal", formula=['"ALARP"'], fill=YELLOW_FILL),
    )
    ws.conditional_formatting.add(
        risk_range,
        CellIsRule(operator="equal", formula=['"Acceptable"'], fill=GREEN_FILL),
    )


def _build_risk_evaluation(wb: Workbook, hazards: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Risk_Evaluation")
    headers = ["Hazard ID", "Risk Level", "Acceptability", "Rationale"]
    widths = [12, 18, 18, 60]
    _style_headers(ws, headers, widths)

    for i, h in enumerate(hazards, start=2):
        _style_cell(ws, i, 1, h["id"])
        # Pull Risk Level from Risk_Estimation sheet via formula
        cell_b = ws.cell(row=i, column=2, value=f"=Risk_Estimation!D{i}")
        cell_b.alignment = CELL_ALIGNMENT
        cell_b.border = THIN_BORDER
        # Acceptability mirrors Risk Level (same formula, or reference)
        cell_c = ws.cell(row=i, column=3, value=f"=Risk_Estimation!D{i}")
        cell_c.alignment = CELL_ALIGNMENT
        cell_c.border = THIN_BORDER
        # Rationale -- only populated for ALARP/Unacceptable in example
        _style_cell(ws, i, 4, h.get("rationale", ""))

    # Conditional formatting on Acceptability (column C)
    acc_range = f"C2:C{len(hazards) + 1}"
    ws.conditional_formatting.add(
        acc_range,
        CellIsRule(operator="equal", formula=['"Unacceptable"'], fill=RED_FILL),
    )
    ws.conditional_formatting.add(
        acc_range,
        CellIsRule(operator="equal", formula=['"ALARP"'], fill=YELLOW_FILL),
    )
    ws.conditional_formatting.add(
        acc_range,
        CellIsRule(operator="equal", formula=['"Acceptable"'], fill=GREEN_FILL),
    )


def _build_risk_controls(wb: Workbook, hazards: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Risk_Controls")
    headers = [
        "Control ID", "Hazard ID", "Control Measure", "Type", "Verification Method",
    ]
    widths = [12, 12, 45, 28, 50]
    _style_headers(ws, headers, widths)

    # Type data validation
    type_dv = DataValidation(
        type="list",
        formula1='"Inherent safety by design,Protective measure,Information for safety"',
    )
    type_dv.error = "Select a valid control type"
    type_dv.errorTitle = "Invalid Type"
    ws.add_data_validation(type_dv)

    for i, h in enumerate(hazards, start=2):
        _style_cell(ws, i, 1, h["rc_id"])
        _style_cell(ws, i, 2, h["id"])
        _style_cell(ws, i, 3, h["control"])
        _style_cell(ws, i, 4, h["control_type"])
        _style_cell(ws, i, 5, h["verification"])
        type_dv.add(ws.cell(row=i, column=4))


def _build_residual_risk(wb: Workbook, hazards: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Residual_Risk")
    headers = [
        "Control ID", "Hazard ID", "Post-Control Severity (S1-S5)",
        "Post-Control Probability (P1-P5)", "Residual Risk Level", "Rationale",
    ]
    widths = [12, 12, 26, 28, 22, 55]
    _style_headers(ws, headers, widths)

    # Data validation
    sev_dv = DataValidation(type="whole", operator="between", formula1="1", formula2="5")
    prob_dv = DataValidation(type="whole", operator="between", formula1="1", formula2="5")
    ws.add_data_validation(sev_dv)
    ws.add_data_validation(prob_dv)

    for i, h in enumerate(hazards, start=2):
        _style_cell(ws, i, 1, h["rc_id"])
        _style_cell(ws, i, 2, h["id"])
        _style_cell(ws, i, 3, h["post_severity"])
        _style_cell(ws, i, 4, h["post_probability"])
        # Residual Risk Level formula (same matrix lookup)
        formula = _risk_level_formula(f"C{i}", f"D{i}")
        cell = ws.cell(row=i, column=5, value=formula)
        cell.alignment = CELL_ALIGNMENT
        cell.border = THIN_BORDER
        # Rationale column
        _style_cell(ws, i, 6, h.get("rationale", ""))

        sev_dv.add(ws.cell(row=i, column=3))
        prob_dv.add(ws.cell(row=i, column=4))

    # Conditional formatting on Residual Risk Level (column E)
    risk_range = f"E2:E{len(hazards) + 1}"
    ws.conditional_formatting.add(
        risk_range,
        CellIsRule(operator="equal", formula=['"Unacceptable"'], fill=RED_FILL),
    )
    ws.conditional_formatting.add(
        risk_range,
        CellIsRule(operator="equal", formula=['"ALARP"'], fill=YELLOW_FILL),
    )
    ws.conditional_formatting.add(
        risk_range,
        CellIsRule(operator="equal", formula=['"Acceptable"'], fill=GREEN_FILL),
    )

    # ALARP rationale enforcement: red fill when risk is ALARP but rationale is empty.
    # NOTE: openpyxl FormulaRule uses relative references anchored to the top-left
    # cell of the applied range.  When the range is F2:F{n}, the formula
    # AND(E2="ALARP",F2="") auto-shifts per row (E3/F3 for row 3, etc.).
    # This is intentional — do NOT use absolute ($E$2) references here.
    rationale_range = f"F2:F{len(hazards) + 1}"
    ws.conditional_formatting.add(
        rationale_range,
        FormulaRule(
            formula=[f'AND(E2="ALARP",F2="")'],
            fill=RED_FILL,
        ),
    )


def _build_overall_residual_risk(wb: Workbook) -> None:
    """Build the Overall_Residual_Risk sheet (ISO 14971 Clause 7)."""
    ws = wb.create_sheet("Overall_Residual_Risk")
    headers = ["Category", "Description", "Status", "Notes"]
    widths = [28, 60, 22, 70]
    _style_headers(ws, headers, widths)

    rows = [
        (
            "Individual Residual Risks",
            "All individual hazard residual risks are documented in the Residual_Risk sheet",
            '=IF(COUNTIF(Residual_Risk!E:E,"Unacceptable")=0,"Acceptable","Review Required")',
            "",
        ),
        (
            "Risk Interaction Analysis",
            "Assessment of whether residual risks interact or compound",
            "Reviewed",
            "HAZ-001 (misclassification) and HAZ-002 (false artifact) could compound: "
            "if artifact flag suppresses real desat AND classification misses it, double failure "
            "path exists. Mitigated by expert queue (RC-005) which catches both.",
        ),
        (
            "Risk Clusters Identified",
            "Groups of related residual risks",
            "Documented",
            "Cluster 1: Information hazards (HAZ-001, HAZ-002, HAZ-005) — all relate to "
            "incorrect clinical information. Cluster 2: Detection hazards (HAZ-003, HAZ-004) "
            "— relate to system configuration/connectivity.",
        ),
        (
            "Benefit-Risk Conclusion",
            "Overall assessment of acceptability considering intended use benefits",
            "Acceptable",
            "Benefits of automated SpO2 triage (reduced alarm fatigue, faster clinical response, "
            "GA-adjusted thresholds) outweigh residual risks. All unacceptable risks reduced to "
            "ALARP or Acceptable. No unacceptable residual risks remain.",
        ),
        (
            "Post-Market Monitoring Plan",
            "How field data feeds back into risk analysis",
            "Defined",
            "Complaint handling per QMS procedure. Quarterly review of classification accuracy "
            "vs. clinical outcomes. CAPA process triggers re-analysis if new failure modes identified.",
        ),
    ]

    for i, (cat, desc, status, notes) in enumerate(rows, start=2):
        _style_cell(ws, i, 1, cat)
        _style_cell(ws, i, 2, desc)
        # Status may be a formula (row 2)
        if str(status).startswith("="):
            cell = ws.cell(row=i, column=3, value=status)
            cell.alignment = CELL_ALIGNMENT
            cell.border = THIN_BORDER
        else:
            _style_cell(ws, i, 3, status)
        _style_cell(ws, i, 4, notes)

    # Summary formula row
    summary_row = len(rows) + 2  # row 7
    _style_cell(ws, summary_row, 1, "Overall Conclusion")
    formula = (
        '=IF(AND(C2="Acceptable",C3="Reviewed",C4="Documented",'
        'C5="Acceptable",C6="Defined"),'
        '"ACCEPTABLE — Proceed to release",'
        '"REVIEW REQUIRED — Resolve open items")'
    )
    cell = ws.cell(row=summary_row, column=2, value=formula)
    cell.alignment = CELL_ALIGNMENT
    cell.border = THIN_BORDER
    cell.font = Font(bold=True, size=11)

    # Conditional formatting on the conclusion cell
    ws.conditional_formatting.add(
        f"B{summary_row}:B{summary_row}",
        CellIsRule(operator="containsText", formula=['"REVIEW REQUIRED"'], fill=RED_FILL),
    )
    ws.conditional_formatting.add(
        f"B{summary_row}:B{summary_row}",
        CellIsRule(operator="containsText", formula=['"ACCEPTABLE"'], fill=GREEN_FILL),
    )


def _build_fmea(wb: Workbook, hazards: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("FMEA")
    headers = [
        "Hazard ID", "Failure Mode", "Effect", "Cause",
        "Severity (S)", "Occurrence (O)", "Detection (D)", "RPN",
    ]
    widths = [12, 45, 40, 45, 14, 14, 14, 10]
    _style_headers(ws, headers, widths)

    # Data validations for S, O, D (1-10)
    for col_idx in (5, 6, 7):
        dv = DataValidation(type="whole", operator="between", formula1="1", formula2="10")
        dv.error = "Value must be between 1 and 10"
        ws.add_data_validation(dv)
        for i in range(len(hazards)):
            dv.add(ws.cell(row=i + 2, column=col_idx))

    for i, h in enumerate(hazards, start=2):
        _style_cell(ws, i, 1, h["id"])
        _style_cell(ws, i, 2, h["fm"])
        _style_cell(ws, i, 3, h["effect"])
        _style_cell(ws, i, 4, h["cause"])
        _style_cell(ws, i, 5, h["fmea_s"])
        _style_cell(ws, i, 6, h["fmea_o"])
        _style_cell(ws, i, 7, h["fmea_d"])
        # RPN formula: =E*F*G
        rpn_cell = ws.cell(row=i, column=8, value=f"=E{i}*F{i}*G{i}")
        rpn_cell.alignment = CELL_ALIGNMENT
        rpn_cell.border = THIN_BORDER

    # Conditional formatting on RPN (column H)
    rpn_range = f"H2:H{len(hazards) + 1}"
    ws.conditional_formatting.add(
        rpn_range,
        CellIsRule(operator="greaterThanOrEqual", formula=["150"], fill=RED_FILL),
    )
    ws.conditional_formatting.add(
        rpn_range,
        CellIsRule(operator="between", formula=["100", "149"], fill=YELLOW_FILL),
    )
    ws.conditional_formatting.add(
        rpn_range,
        CellIsRule(operator="lessThan", formula=["100"], fill=GREEN_FILL),
    )


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def generate(device_name: str, hazards: list[dict[str, Any]]) -> str:
    """Generate the risk analysis XLSX and return the output path."""
    wb = Workbook()
    # Rename the default sheet to Document_Control
    doc_ws = wb.active
    doc_ws.title = "Document_Control"
    _build_document_control(doc_ws, device_name)

    _build_hazard_identification(wb, hazards)
    _build_risk_estimation(wb, hazards)
    _build_risk_evaluation(wb, hazards)
    _build_risk_controls(wb, hazards)
    _build_residual_risk(wb, hazards)
    _build_overall_residual_risk(wb)
    _build_fmea(wb, hazards)

    # Sanitize device name for filename
    safe_name = re.sub(r"[^a-zA-Z0-9_-]", "-", device_name.lower().strip())
    safe_name = re.sub(r"-+", "-", safe_name).strip("-")

    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_dir = os.path.join(os.path.dirname(script_dir), "output")
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, f"risk-analysis-{safe_name}.xlsx")

    wb.save(output_path)
    return output_path


def _blank_hazards() -> list[dict[str, Any]]:
    """Return a single blank hazard row as a template."""
    return [
        {
            "id": "HAZ-001",
            "hazard": "",
            "situation": "",
            "harm": "",
            "sequence": "",
            "severity": 1, "probability": 1,
            "rc_id": "RC-001",
            "control": "",
            "control_type": "Inherent safety by design",
            "verification": "",
            "post_severity": 1, "post_probability": 1,
            "rationale": "",
            "fm": "",
            "effect": "",
            "cause": "",
            "fmea_s": 1, "fmea_o": 1, "fmea_d": 1,
        }
    ]


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Generate ISO 14971 risk management XLSX"
    )
    parser.add_argument(
        "--device-name",
        type=str,
        default=None,
        help="Name of the medical device (used in filename)",
    )
    parser.add_argument(
        "--example",
        type=str,
        choices=["spo2"],
        default=None,
        help="Generate with built-in example data (spo2)",
    )
    args = parser.parse_args()

    if args.example == "spo2":
        device_name = "SpO2 AI Eval Pipeline"
        hazards = _spo2_hazards()
    elif args.device_name:
        device_name = args.device_name
        hazards = _blank_hazards()
    else:
        parser.error("Provide --example spo2 or --device-name 'Device Name'")
        return  # unreachable but satisfies type checker

    output_path = generate(device_name, hazards)
    print(f"Risk analysis generated: {output_path}")
    print(f"Sheets: 8 (Document_Control, Hazard_Identification, Risk_Estimation, "
          f"Risk_Evaluation, Risk_Controls, Residual_Risk, Overall_Residual_Risk, FMEA)")
    print(f"Hazards: {len(hazards)}")


if __name__ == "__main__":
    main()
