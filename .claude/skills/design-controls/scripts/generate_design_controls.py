#!/usr/bin/env python3
"""Generate IEC 62304 / ISO 13485 design controls traceability matrices as XLSX.

Standalone script — only requires openpyxl.

Usage:
    python scripts/generate_design_controls.py --example spo2
    python scripts/generate_design_controls.py --device-name "My Device" --safety-class C
"""

import argparse
import os
import re
from datetime import date
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGN = Alignment(vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Conditional formatting fills
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
GREEN_FONT = Font(color="006100")
RED_FONT = Font(color="9C0006")
YELLOW_FONT = Font(color="9C6500")


# ---------------------------------------------------------------------------
# SpO2 example data
# ---------------------------------------------------------------------------
def get_spo2_example() -> dict[str, Any]:
    """Return hardcoded SpO2 AI Eval Pipeline example data."""
    device_name = "SpO2 AI Eval Pipeline"
    safety_class = "C"

    user_needs = [
        {
            "id": "UN-001",
            "statement": "System shall classify SpO2 traces into urgency tiers (normal, borderline, urgent, emergency)",
            "source": "Clinical Requirements",
            "priority": "High",
            "acceptance_criteria": "Classification accuracy >= 85% across all tiers on held-out test set",
        },
        {
            "id": "UN-002",
            "statement": "System shall detect desaturation events below GA-adjusted thresholds",
            "source": "Neonatal Guidelines (Castillo 2008)",
            "priority": "High",
            "acceptance_criteria": "Sensitivity >= 90% for desaturation events; false negative rate < 5%",
        },
        {
            "id": "UN-003",
            "statement": "System shall generate structured handoff reports for clinical review",
            "source": "Nurse Workflow Analysis",
            "priority": "Medium",
            "acceptance_criteria": "Reports contain all required fields per SBAR format; 95% nurse satisfaction",
        },
        {
            "id": "UN-004",
            "statement": "System shall flag motion artifacts using accelerometer correlation",
            "source": "Signal Processing Requirements",
            "priority": "Medium",
            "acceptance_criteria": "Artifact detection precision >= 80%; recall >= 75%",
        },
        {
            "id": "UN-005",
            "statement": "System shall route ambiguous cases to expert queue with confidence scores",
            "source": "Clinical Safety Assessment",
            "priority": "High",
            "acceptance_criteria": "All cases with confidence < 0.7 routed to expert queue; zero missed critical cases",
        },
    ]

    design_inputs = [
        {
            "id": "DI-001",
            "linked_un": "UN-001",
            "requirement": "Implement 4-tier classification (normal >= 95%, borderline 90-94%, urgent 85-89%, emergency < 85%) with GA-adjusted thresholds",
            "type": "Functional",
            "safety_class": safety_class,
            "safety_class_rationale": "Class C: Controls desaturation detection; misclassification could delay clinical intervention causing serious harm",
        },
        {
            "id": "DI-002",
            "linked_un": "UN-002",
            "requirement": "Detect SpO2 drops >= 4% from baseline sustained for >= 10 seconds, adjusted by gestational age per Castillo 2008 reference ranges",
            "type": "Performance",
            "safety_class": safety_class,
            "safety_class_rationale": "Class C: Desat detection failure could result in undetected hypoxemia",
        },
        {
            "id": "DI-003",
            "linked_un": "UN-003",
            "requirement": "Generate JSON-structured handoff reports including patient ID, tier classification, desaturation events, trend summary, and recommended actions",
            "type": "Functional",
            "safety_class": "B",
            "safety_class_rationale": "Class B: Report generation supports clinical workflow; failure delays but does not prevent intervention (alarms still active)",
        },
        {
            "id": "DI-004",
            "linked_un": "UN-004",
            "requirement": "Correlate accelerometer signal with SpO2 waveform; flag segments where motion correlation coefficient > 0.6 as artifact-contaminated",
            "type": "Performance",
            "safety_class": "B",
            "safety_class_rationale": "Class B: Artifact flagging is supportive; false artifact flag on real data is caught by expert review queue",
        },
        {
            "id": "DI-005",
            "linked_un": "UN-005",
            "requirement": "Calculate classification confidence score (0.0-1.0) per trace; route to expert queue when confidence < 0.7; include top 3 differential classifications",
            "type": "Safety",
            "safety_class": safety_class,
            "safety_class_rationale": "Class C: Expert queue routing failure could suppress ambiguous cases requiring clinical judgment",
        },
    ]

    design_outputs = [
        {
            "id": "DO-001",
            "linked_di": "DI-001",
            "description": "Tier classification module with GA-adjusted threshold lookup table and 4-class output enum",
            "verification_method": "Test",
        },
        {
            "id": "DO-002",
            "linked_di": "DI-002",
            "description": "Desaturation detection algorithm with configurable baseline window and duration filter",
            "verification_method": "Test",
        },
        {
            "id": "DO-003",
            "linked_di": "DI-003",
            "description": "Handoff report generator producing SBAR-format JSON with all required clinical fields",
            "verification_method": "Inspection",
        },
        {
            "id": "DO-004",
            "linked_di": "DI-004",
            "description": "Motion artifact detector using sliding-window cross-correlation with accelerometer channel",
            "verification_method": "Test",
        },
        {
            "id": "DO-005",
            "linked_di": "DI-005",
            "description": "Confidence scoring module with softmax-based certainty metric and expert routing logic",
            "verification_method": "Test",
        },
    ]

    verification = [
        {
            "id": "VER-001",
            "linked_do": "DO-001",
            "test_protocol": "TP-001: Run classifier on 500-sample test set with known tier labels",
            "pass_criteria": "Overall accuracy >= 85%; per-tier F1 >= 0.80",
            "status": "Pending",
            "test_data_source": "Synthetic dataset, N=300 traces, algorithm-generated per GA category",
        },
        {
            "id": "VER-002",
            "linked_do": "DO-002",
            "test_protocol": "TP-002: Inject synthetic desaturation events at known timestamps into 100 traces",
            "pass_criteria": "Sensitivity >= 90%; false negative rate < 5%",
            "status": "Pending",
            "test_data_source": "Internal NICU reference dataset, N=500 annotated events (Castillo 2008 thresholds)",
        },
        {
            "id": "VER-003",
            "linked_do": "DO-003",
            "test_protocol": "TP-003: Validate 50 generated reports against SBAR field schema",
            "pass_criteria": "100% schema compliance; all required fields present",
            "status": "Pending",
            "test_data_source": "Synthetic handoff reports, N=50, template-validated",
        },
        {
            "id": "VER-004",
            "linked_do": "DO-004",
            "test_protocol": "TP-004: Run artifact detector on 200 segments with labeled motion/no-motion ground truth",
            "pass_criteria": "Precision >= 80%; recall >= 75%",
            "status": "Pending",
            "test_data_source": "Synthetic motion artifact dataset, N=200 traces with accelerometer correlation",
        },
        {
            "id": "VER-005",
            "linked_do": "DO-005",
            "test_protocol": "TP-005: Verify routing logic on 300 cases spanning confidence range 0.0-1.0",
            "pass_criteria": "100% of cases with confidence < 0.7 routed; zero missed critical (emergency tier)",
            "status": "Pending",
            "test_data_source": "Synthetic ambiguous cases, N=100, confidence score distribution validated",
        },
    ]

    validation = [
        {
            "id": "VAL-001",
            "linked_un": "UN-001",
            "validation_method": "Clinical simulation with 3 neonatologists reviewing 100 classified traces",
            "acceptance_criteria": "Clinician agreement with system classification >= 85%",
            "status": "Pending",
        },
        {
            "id": "VAL-002",
            "linked_un": "UN-002",
            "validation_method": "Retrospective analysis on NICU dataset with chart-confirmed desaturation events",
            "acceptance_criteria": "Sensitivity >= 90% vs. chart-confirmed events",
            "status": "Pending",
        },
        {
            "id": "VAL-003",
            "linked_un": "UN-003",
            "validation_method": "Usability study: 10 NICU nurses use handoff reports in simulated shift change",
            "acceptance_criteria": ">= 95% task completion; SUS score >= 70",
            "status": "Pending",
        },
        {
            "id": "VAL-004",
            "linked_un": "UN-004",
            "validation_method": "Bench test with motion simulator generating known artifact patterns",
            "acceptance_criteria": "Artifact detection aligns with expert annotation in >= 80% of segments",
            "status": "Pending",
        },
        {
            "id": "VAL-005",
            "linked_un": "UN-005",
            "validation_method": "Prospective pilot: expert queue reviewed by senior clinician over 2-week period",
            "acceptance_criteria": "Zero false negatives for emergency-tier cases; expert queue volume < 20% of total",
            "status": "Pending",
        },
    ]

    return {
        "device_name": device_name,
        "safety_class": safety_class,
        "user_needs": user_needs,
        "design_inputs": design_inputs,
        "design_outputs": design_outputs,
        "verification": verification,
        "validation": validation,
    }


# ---------------------------------------------------------------------------
# Blank scaffold (for custom --device-name usage)
# ---------------------------------------------------------------------------
def get_blank_scaffold(device_name: str, safety_class: str) -> dict[str, Any]:
    """Return a minimal scaffold with one row per sheet for a custom device."""
    return {
        "device_name": device_name,
        "safety_class": safety_class,
        "user_needs": [
            {
                "id": "UN-001",
                "statement": "<Enter user need statement>",
                "source": "<Source>",
                "priority": "High",
                "acceptance_criteria": "<Measurable acceptance criteria>",
            }
        ],
        "design_inputs": [
            {
                "id": "DI-001",
                "linked_un": "UN-001",
                "requirement": "<Testable requirement derived from UN-001>",
                "type": "Functional",
                "safety_class": safety_class,
                "safety_class_rationale": "<Justification for assigned safety class>",
            }
        ],
        "design_outputs": [
            {
                "id": "DO-001",
                "linked_di": "DI-001",
                "description": "<Output that satisfies DI-001>",
                "verification_method": "Test",
            }
        ],
        "verification": [
            {
                "id": "VER-001",
                "linked_do": "DO-001",
                "test_protocol": "<Test procedure reference>",
                "pass_criteria": "<Quantitative pass threshold>",
                "status": "Pending",
                "test_data_source": "<Description of test data origin and size>",
            }
        ],
        "validation": [
            {
                "id": "VAL-001",
                "linked_un": "UN-001",
                "validation_method": "<Validation approach>",
                "acceptance_criteria": "<Evidence threshold>",
                "status": "Pending",
            }
        ],
    }


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------
def style_header_row(ws, headers: list[str]) -> None:
    """Write and style the header row."""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def auto_width(ws, min_width: int = 12, max_width: int = 60) -> None:
    """Auto-adjust column widths based on content."""
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            if cell.value:
                # Handle multiline — use longest line
                lines = str(cell.value).split("\n")
                line_max = max(len(line) for line in lines)
                max_len = max(max_len, line_max)
        adjusted = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


def add_status_validation(ws, col_letter: str, row_start: int, row_end: int) -> None:
    """Add Pass/Fail/Pending dropdown validation to a status column."""
    dv = DataValidation(
        type="list",
        formula1='"Pass,Fail,Pending"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Invalid Status",
        error="Status must be Pass, Fail, or Pending.",
    )
    dv.sqref = f"{col_letter}{row_start}:{col_letter}{row_end}"
    ws.add_data_validation(dv)


def add_status_conditional_formatting(ws, col_letter: str, row_start: int, row_end: int) -> None:
    """Add green/red/yellow conditional formatting to a status column."""
    cell_range = f"{col_letter}{row_start}:{col_letter}{row_end}"
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="equal", formula=['"Pass"'], fill=GREEN_FILL, font=GREEN_FONT),
    )
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="equal", formula=['"Fail"'], fill=RED_FILL, font=RED_FONT),
    )
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="equal", formula=['"Pending"'], fill=YELLOW_FILL, font=YELLOW_FONT),
    )


def style_data_rows(ws, start_row: int, end_row: int, num_cols: int) -> None:
    """Apply alignment and borders to data rows."""
    for row in range(start_row, end_row + 1):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = CELL_ALIGN
            cell.border = THIN_BORDER


def build_document_control(wb: Workbook, data: dict[str, Any]) -> None:
    """Build the Document_Control sheet as the first sheet in the workbook.

    Uses wb.active (the default sheet created by Workbook()) and renames it
    so it stays at tab index 0.
    """
    ws = wb.active
    ws.title = "Document_Control"

    # Row 1: Title merged across A:D
    ws.merge_cells("A1:D1")
    title_cell = ws.cell(row=1, column=1, value="Document Control")
    title_cell.font = HEADER_FONT
    title_cell.fill = HEADER_FILL
    title_cell.alignment = HEADER_ALIGN
    title_cell.border = THIN_BORDER
    # Style merged area borders
    for col in range(2, 5):
        c = ws.cell(row=1, column=col)
        c.border = THIN_BORDER
        c.fill = HEADER_FILL

    device_name = data["device_name"]
    safety_class = data["safety_class"]

    # Rows 3-7: Metadata key-value pairs
    metadata = [
        ("Document Title", f"Design Controls \u2014 {device_name}"),
        ("Document Version", "1.0"),
        ("Device/Software Version", f"{device_name} v2.1"),
        ("Date Generated", date.today().strftime("%Y-%m-%d")),
        ("Safety Class", safety_class),
    ]
    for idx, (label, value) in enumerate(metadata):
        row = 3 + idx
        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font = Font(bold=True, size=11)
        label_cell.alignment = CELL_ALIGN
        label_cell.border = THIN_BORDER
        value_cell = ws.cell(row=row, column=2, value=value)
        value_cell.alignment = CELL_ALIGN
        value_cell.border = THIN_BORDER

    # Rows 8-10: QMS integration fields
    qms_fields = [
        ("QMS Document Number", ""),
        ("SOP Reference", ""),
        ("Related Risk Management File", ""),
    ]
    for idx, (label, value) in enumerate(qms_fields):
        row = 8 + idx
        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font = Font(bold=True, size=11)
        label_cell.alignment = CELL_ALIGN
        label_cell.border = THIN_BORDER
        value_cell = ws.cell(row=row, column=2, value=value)
        value_cell.alignment = CELL_ALIGN
        value_cell.border = THIN_BORDER

    # Row 12: Approval Signatures header (merged across A:D)
    ws.merge_cells("A12:D12")
    sig_cell = ws.cell(row=12, column=1, value="Approval Signatures")
    sig_cell.font = Font(bold=True, size=11)
    sig_cell.alignment = CELL_ALIGN
    sig_cell.border = THIN_BORDER
    for col in range(2, 5):
        c = ws.cell(row=12, column=col)
        c.border = THIN_BORDER

    # Row 13: Column headers for signature table
    sig_headers = ["Role", "Name", "Date", "Signature"]
    for col_idx, header in enumerate(sig_headers, 1):
        cell = ws.cell(row=13, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    # Rows 14-16: Signature rows
    sig_roles = ["Prepared By", "Reviewed By (QA)", "Approved By (Management)"]
    for idx, role in enumerate(sig_roles):
        row = 14 + idx
        ws.cell(row=row, column=1, value=role).border = THIN_BORDER
        ws.cell(row=row, column=1).alignment = CELL_ALIGN
        for col in range(2, 5):
            c = ws.cell(row=row, column=col, value="")
            c.border = THIN_BORDER
            c.alignment = CELL_ALIGN

    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 20


def build_user_needs(wb: Workbook, data: list[dict]) -> None:
    """Build the User_Needs sheet."""
    ws = wb.create_sheet("User_Needs")
    headers = ["ID", "Need Statement", "Source", "Priority", "Acceptance Criteria"]
    style_header_row(ws, headers)

    for i, un in enumerate(data, 2):
        ws.cell(row=i, column=1, value=un["id"])
        ws.cell(row=i, column=2, value=un["statement"])
        ws.cell(row=i, column=3, value=un["source"])
        ws.cell(row=i, column=4, value=un["priority"])
        ws.cell(row=i, column=5, value=un["acceptance_criteria"])

    end_row = len(data) + 1
    style_data_rows(ws, 2, end_row, len(headers))
    auto_width(ws)


def build_design_inputs(wb: Workbook, data: list[dict]) -> None:
    """Build the Design_Inputs sheet."""
    ws = wb.create_sheet("Design_Inputs")
    headers = ["ID", "Linked UN", "Requirement", "Type", "SW Safety Class", "Safety Class Rationale"]
    style_header_row(ws, headers)

    for i, di in enumerate(data, 2):
        ws.cell(row=i, column=1, value=di["id"])
        ws.cell(row=i, column=2, value=di["linked_un"])
        ws.cell(row=i, column=3, value=di["requirement"])
        ws.cell(row=i, column=4, value=di["type"])
        ws.cell(row=i, column=5, value=di["safety_class"])
        ws.cell(row=i, column=6, value=di.get("safety_class_rationale", ""))

    end_row = len(data) + 1
    style_data_rows(ws, 2, end_row, len(headers))
    auto_width(ws)


def build_design_outputs(wb: Workbook, data: list[dict]) -> None:
    """Build the Design_Outputs sheet."""
    ws = wb.create_sheet("Design_Outputs")
    headers = ["ID", "Linked DI", "Output Description", "Verification Method"]
    style_header_row(ws, headers)

    for i, do in enumerate(data, 2):
        ws.cell(row=i, column=1, value=do["id"])
        ws.cell(row=i, column=2, value=do["linked_di"])
        ws.cell(row=i, column=3, value=do["description"])
        ws.cell(row=i, column=4, value=do["verification_method"])

    end_row = len(data) + 1
    style_data_rows(ws, 2, end_row, len(headers))
    auto_width(ws)


def build_verification(wb: Workbook, data: list[dict]) -> None:
    """Build the Verification sheet."""
    ws = wb.create_sheet("Verification")
    headers = ["ID", "Linked DO", "Test Protocol", "Pass Criteria", "Status", "Test Data Source"]
    style_header_row(ws, headers)

    for i, ver in enumerate(data, 2):
        ws.cell(row=i, column=1, value=ver["id"])
        ws.cell(row=i, column=2, value=ver["linked_do"])
        ws.cell(row=i, column=3, value=ver["test_protocol"])
        ws.cell(row=i, column=4, value=ver["pass_criteria"])
        ws.cell(row=i, column=5, value=ver["status"])
        ws.cell(row=i, column=6, value=ver.get("test_data_source", ""))

    end_row = len(data) + 1
    style_data_rows(ws, 2, end_row, len(headers))

    # Status column = E
    status_col = "E"
    add_status_validation(ws, status_col, 2, end_row)
    add_status_conditional_formatting(ws, status_col, 2, end_row)
    auto_width(ws)


def build_validation(wb: Workbook, data: list[dict]) -> None:
    """Build the Validation sheet."""
    ws = wb.create_sheet("Validation")
    headers = ["ID", "Linked UN", "Validation Method", "Acceptance Criteria", "Status"]
    style_header_row(ws, headers)

    for i, val in enumerate(data, 2):
        ws.cell(row=i, column=1, value=val["id"])
        ws.cell(row=i, column=2, value=val["linked_un"])
        ws.cell(row=i, column=3, value=val["validation_method"])
        ws.cell(row=i, column=4, value=val["acceptance_criteria"])
        ws.cell(row=i, column=5, value=val["status"])

    end_row = len(data) + 1
    style_data_rows(ws, 2, end_row, len(headers))

    # Status column = E
    status_col = "E"
    add_status_validation(ws, status_col, 2, end_row)
    add_status_conditional_formatting(ws, status_col, 2, end_row)
    auto_width(ws)


def build_traceability_matrix(wb: Workbook, data: dict[str, Any]) -> None:
    """Build the Traceability_Matrix sheet with VLOOKUP formulas."""
    ws = wb.create_sheet("Traceability_Matrix")
    headers = [
        "UN ID",
        "User Need",
        "DI ID",
        "Design Input",
        "DO ID",
        "Design Output",
        "VER ID",
        "VER Status",
        "VAL ID",
        "VAL Status",
    ]
    style_header_row(ws, headers)

    num_rows = len(data["user_needs"])

    for i, un in enumerate(data["user_needs"], 2):
        un_id = un["id"]
        # Find linked DI, DO, VER, VAL by tracing the chain
        di = next((d for d in data["design_inputs"] if d["linked_un"] == un_id), None)
        do = next((d for d in data["design_outputs"] if di and d["linked_di"] == di["id"]), None)
        ver = next((v for v in data["verification"] if do and v["linked_do"] == do["id"]), None)
        val = next((v for v in data["validation"] if v["linked_un"] == un_id), None)

        di_id = di["id"] if di else ""
        do_id = do["id"] if do else ""
        ver_id = ver["id"] if ver else ""
        val_id = val["id"] if val else ""

        # Column A: UN ID (hardcoded for lookup reference)
        ws.cell(row=i, column=1, value=un_id)

        # Column B: VLOOKUP to pull Need Statement from User_Needs
        ws.cell(
            row=i, column=2,
            value=f'=VLOOKUP(A{i},User_Needs!A:E,2,FALSE)',
        )

        # Column C: DI ID
        ws.cell(row=i, column=3, value=di_id)

        # Column D: VLOOKUP to pull Requirement from Design_Inputs
        if di_id:
            ws.cell(
                row=i, column=4,
                value=f'=VLOOKUP(C{i},Design_Inputs!A:F,3,FALSE)',
            )
        else:
            ws.cell(row=i, column=4, value="")

        # Column E: DO ID
        ws.cell(row=i, column=5, value=do_id)

        # Column F: VLOOKUP to pull Output Description from Design_Outputs
        if do_id:
            ws.cell(
                row=i, column=6,
                value=f'=VLOOKUP(E{i},Design_Outputs!A:D,3,FALSE)',
            )
        else:
            ws.cell(row=i, column=6, value="")

        # Column G: VER ID
        ws.cell(row=i, column=7, value=ver_id)

        # Column H: VLOOKUP to pull VER Status from Verification
        if ver_id:
            ws.cell(
                row=i, column=8,
                value=f'=VLOOKUP(G{i},Verification!A:F,5,FALSE)',
            )
        else:
            ws.cell(row=i, column=8, value="")

        # Column I: VAL ID
        ws.cell(row=i, column=9, value=val_id)

        # Column J: VLOOKUP to pull VAL Status from Validation
        if val_id:
            ws.cell(
                row=i, column=10,
                value=f'=VLOOKUP(I{i},Validation!A:E,5,FALSE)',
            )
        else:
            ws.cell(row=i, column=10, value="")

    end_row = num_rows + 1
    style_data_rows(ws, 2, end_row, len(headers))

    # Add conditional formatting on VER Status (H) and VAL Status (J)
    for col_letter in ["H", "J"]:
        add_status_conditional_formatting(ws, col_letter, 2, end_row)

    # Summary stats section below the matrix
    summary_start = end_row + 2
    ws.cell(row=summary_start, column=1, value="Summary Statistics")
    ws.cell(row=summary_start, column=1).font = Font(bold=True, size=12)

    # Verification stats
    ws.cell(row=summary_start + 1, column=1, value="Verification")
    ws.cell(row=summary_start + 1, column=1).font = Font(bold=True)
    ws.cell(row=summary_start + 1, column=2, value="Pass:")
    ws.cell(row=summary_start + 1, column=3, value='=COUNTIF(Verification!E:E,"Pass")')
    ws.cell(row=summary_start + 1, column=4, value="Fail:")
    ws.cell(row=summary_start + 1, column=5, value='=COUNTIF(Verification!E:E,"Fail")')
    ws.cell(row=summary_start + 1, column=6, value="Pending:")
    ws.cell(row=summary_start + 1, column=7, value='=COUNTIF(Verification!E:E,"Pending")')

    # Validation stats
    ws.cell(row=summary_start + 2, column=1, value="Validation")
    ws.cell(row=summary_start + 2, column=1).font = Font(bold=True)
    ws.cell(row=summary_start + 2, column=2, value="Pass:")
    ws.cell(row=summary_start + 2, column=3, value='=COUNTIF(Validation!E:E,"Pass")')
    ws.cell(row=summary_start + 2, column=4, value="Fail:")
    ws.cell(row=summary_start + 2, column=5, value='=COUNTIF(Validation!E:E,"Fail")')
    ws.cell(row=summary_start + 2, column=6, value="Pending:")
    ws.cell(row=summary_start + 2, column=7, value='=COUNTIF(Validation!E:E,"Pending")')

    auto_width(ws)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def slugify(name: str) -> str:
    """Convert device name to filename-safe slug."""
    slug = name.lower().strip()
    slug = re.sub(r"[^a-z0-9]+", "-", slug)
    slug = slug.strip("-")
    return slug


def generate(data: dict[str, Any]) -> str:
    """Generate the XLSX workbook and return the output path."""
    wb = Workbook()

    # Build Document_Control first (uses wb.active, stays at tab index 0)
    build_document_control(wb, data)

    # Build all 6 data sheets
    build_user_needs(wb, data["user_needs"])
    build_design_inputs(wb, data["design_inputs"])
    build_design_outputs(wb, data["design_outputs"])
    build_verification(wb, data["verification"])
    build_validation(wb, data["validation"])
    build_traceability_matrix(wb, data)

    # Set Document_Control as the active/first-visible sheet
    wb.active = 0

    # Save
    slug = slugify(data["device_name"])
    output_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "output")
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, f"design-controls-{slug}.xlsx")
    wb.save(output_path)
    return output_path


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Generate IEC 62304 / ISO 13485 design controls traceability matrix (XLSX)."
    )
    parser.add_argument(
        "--example",
        choices=["spo2"],
        help="Generate a pre-populated example (spo2 = SpO2 AI Eval Pipeline).",
    )
    parser.add_argument(
        "--device-name",
        type=str,
        help="Device/product name for the header.",
    )
    parser.add_argument(
        "--safety-class",
        choices=["A", "B", "C"],
        default="C",
        help="IEC 62304 SW safety class (default: C).",
    )
    args = parser.parse_args()

    if args.example == "spo2":
        data = get_spo2_example()
    elif args.device_name:
        data = get_blank_scaffold(args.device_name, args.safety_class)
    else:
        parser.error("Provide --example spo2 or --device-name 'Your Device'.")

    output_path = generate(data)
    print(f"Generated: {output_path}")
    print(f"  Sheets: Document_Control, User_Needs, Design_Inputs, Design_Outputs, Verification, Validation, Traceability_Matrix")
    print(f"  Safety Class: {data['safety_class']}")
    print(f"  Records: {len(data['user_needs'])} user needs → full traceability chain")


if __name__ == "__main__":
    main()
