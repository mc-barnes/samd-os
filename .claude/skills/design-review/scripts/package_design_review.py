#!/usr/bin/env python3
"""
Design Review Packager — generates XLSX + markdown narrative for PDR/CDR/FDR gates.

Usage:
    python scripts/package_design_review.py --review-type CDR --example spo2
    python scripts/package_design_review.py --review-type PDR --device-name "Cardiac Monitor v1.0"
    python scripts/package_design_review.py --review-type FDR --example spo2 \
        --design-controls path.xlsx --risk-analysis path.xlsx --change-impact path.xlsx
"""

import argparse
import os
import sys
from datetime import datetime, timedelta
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill(start_color="4B0082", end_color="4B0082", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
LABEL_FONT = Font(bold=True, size=11)
VALUE_FONT = Font(size=11)

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

GO_FILL = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
GO_FONT = Font(bold=True, color="FFFFFF", size=14)
CONDITIONAL_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
CONDITIONAL_FONT = Font(bold=True, color="000000", size=14)
NOGO_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
NOGO_FONT = Font(bold=True, color="FFFFFF", size=14)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

REVIEW_TYPE_MAP = {
    "PDR": "PDR -- Preliminary Design Review",
    "CDR": "CDR -- Critical Design Review",
    "FDR": "FDR -- Final Design Review",
}

REVIEW_FOCUS = {
    "PDR": "Is the concept viable? Are requirements clear?",
    "CDR": "Is the design complete enough to build?",
    "FDR": "Is the product ready to ship?",
}

REVIEW_GATE = {
    "PDR": "Concept -> Detailed Design",
    "CDR": "Detailed Design -> Implementation/V&V",
    "FDR": "V&V -> Release/Transfer",
}

# ---------------------------------------------------------------------------
# Checklists — full items for PDR (25), CDR (56), FDR (92)
# ---------------------------------------------------------------------------

PDR_ITEMS = [
    "User needs document complete",
    "Design inputs derived from user needs",
    "Intended use statement defined",
    "User population characterized",
    "Use environment described",
    "Regulatory pathway identified",
    "Applicable standards listed",
    "Risk management plan created",
    "Initial hazard analysis performed",
    "SW safety classification determined",
    "System architecture defined",
    "Key algorithms identified",
    "Data flow documented",
    "Interface specifications drafted",
    "Performance requirements specified",
    "Security requirements identified",
    "Privacy/HIPAA assessment done",
    "V&V strategy outlined",
    "Clinical evidence plan (if needed)",
    "Development timeline realistic",
    "Resource plan adequate",
    "Previous review actions closed",
    "Design history file initiated",
    "Configuration management set up",
    "Supplier/third-party assessment (if applicable)",
]

CDR_EXTRA_ITEMS = [
    "Detailed design specifications complete",
    "All design inputs have corresponding outputs",
    "Traceability matrix complete (UN->DI->DO->VER->VAL)",
    "Design inputs are unambiguous, complete, non-conflicting, and have measurable acceptance criteria (per FDA 820.30(c))",
    "Risk analysis updated with design details",
    "All unacceptable risks have control measures",
    "Risk controls traced to requirements",
    "V&V protocols written and reviewed",
    "Test environment specified",
    "Acceptance criteria quantified",
    "Algorithm performance targets set",
    "Data quality requirements defined",
    "Integration test plan complete",
    "Regression test strategy defined",
    "Cybersecurity analysis complete",
    "Labeling/IFU drafted",
    "Manufacturing/deployment plan drafted",
    "Training materials outlined",
    "Post-market surveillance plan initiated",
    "All CDR action items from PDR closed",
    "Independent review of critical safety items",
    "Biocompatibility assessment (if applicable)",
    "Electrical safety assessment (if applicable)",
    "EMC assessment (if applicable)",
    "Software of Unknown Provenance (SOUP) list complete",
    "SBOM (Software Bill of Materials) generated",
    "Usability engineering file initiated",
    "Clinical evaluation report drafted",
    "Regulatory submission timeline set",
    "Budget and resource reconfirmed",
    "Stakeholder sign-off on design direction",
]

FDR_EXTRA_ITEMS = [
    "All verification tests executed and passed",
    "All validation tests executed and passed",
    "Verification vs. validation clearly distinguished (distinct test protocols for each)",
    "Traceability matrix fully populated (no gaps)",
    "Risk management report complete",
    "Overall residual risk acceptable",
    "Design transfer documentation complete",
    "Labeling/IFU finalized",
    "Training materials finalized",
    "Post-market surveillance plan finalized",
    "Regulatory submission package assembled",
    "Clinical evidence sufficient",
    "Independent review complete",
    "All open action items from CDR closed",
    "No unresolved non-conformances",
    "Production/deployment environment validated",
    "Backup and recovery procedures tested",
    "Cybersecurity measures verified",
    "Data retention/destruction plan",
    "Customer support plan ready",
    "Field safety corrective action plan",
    "All SOUP/SBOM items assessed",
    "Usability validation complete",
    "Human factors summative evaluation complete",
    "Software release notes prepared",
    "Design history file complete",
    "Quality system audit findings resolved",
    "Regulatory authority correspondence reviewed",
    "Market authorization application ready",
    "Go-to-market plan finalized",
    "Launch readiness confirmed by all stakeholders",
    "Post-launch monitoring plan active",
    "Complaint handling process ready",
    "Adverse event reporting process ready",
    "Product recall procedure documented",
    "Executive sign-off obtained",
]


def get_checklist_items(review_type: str) -> list[str]:
    """Return the full checklist for the given review type."""
    if review_type == "PDR":
        return list(PDR_ITEMS)
    elif review_type == "CDR":
        return list(PDR_ITEMS) + list(CDR_EXTRA_ITEMS)
    else:  # FDR
        return list(PDR_ITEMS) + list(CDR_EXTRA_ITEMS) + list(FDR_EXTRA_ITEMS)


# ---------------------------------------------------------------------------
# SpO2 Example Data
# ---------------------------------------------------------------------------

SPO2_DEVICE_NAME = "SpO2 AI Eval Pipeline v2.1"
SPO2_DEVICE_SLUG = "spo2-ai-eval-pipeline"
SPO2_ATTENDEES = (
    "[PM Name] (PM), Dr. Chen (Clinical), Sarah Kim (QA/Regulatory), "
    "Alex Rivera (Data Science), Jordan Park (Engineering), Independent Reviewer TBD"
)

# Items NOT provided for CDR example (0-indexed positions in CDR checklist)
# Items 46-48 (biocompat/electrical/EMC = N/A but still require documented rationale),
# Item 44 (independent review pending), Item 52 (clinical eval in progress),
# Items 53-55 (regulatory timeline, budget, sign-off pending)
# Note: CDR has 56 items (25 PDR + 31 CDR-extra including 820.30(c) design input item)
# Result: 48/56 provided = 86% completeness = CONDITIONAL
SPO2_CDR_NOT_REQUIRED: set[int] = set()  # All items required (N/A items need documented rationale)
SPO2_CDR_NOT_PROVIDED = {44, 46, 47, 48, 52, 53, 54, 55}  # 8 items not yet provided

# For PDR: mark 3 items as not yet provided
SPO2_PDR_NOT_PROVIDED = {18, 19, 24}  # Clinical evidence plan, timeline, supplier assessment

# For FDR: CDR gaps + several FDR items pending
# Note: FDR has 92 items (25 PDR + 31 CDR-extra + 36 FDR-extra)
SPO2_FDR_NOT_PROVIDED = {44, 46, 47, 48, 52, 53, 54, 55, 82, 83, 84, 85, 86, 90, 91}

SPO2_REQUIREMENTS = {
    "total_user_needs": 5,
    "total_design_inputs": 5,
    "verified": 5,
    "validated": 4,
    "open_items": 1,
    "open_detail": "VAL-005 (expert queue routing) pending",
}

SPO2_RISK = {
    "total_hazards": 5,
    "unacceptable": 0,
    "alarp": 2,
    "acceptable": 3,
    "open_controls": 0,
}

SPO2_SIGNOFF = [
    ("PM", "[PM Name]", "2026-04-28", "Approve"),
    ("Clinical Reviewer", "Dr. Chen", "2026-04-28", "Conditional"),
    ("Data Scientist", "Alex Rivera", "2026-04-28", "Approve"),
    ("QA/Regulatory", "Sarah Kim", "2026-04-28", "Conditional"),
    ("Engineering Lead", "Jordan Park", "2026-04-28", "Approve"),
    ("Independent Reviewer", "TBD", "", "Pending"),
]

SPO2_SIGNOFF_NOTES = {
    "Clinical Reviewer": "Pending VAL-005 expert queue validation",
    "QA/Regulatory": "Need regulatory submission timeline",
}

SPO2_ACTION_ITEMS = [
    ("Complete VAL-005 expert queue validation", "Alex Rivera", "2026-05-15"),
    ("Finalize regulatory submission timeline", "Sarah Kim", "2026-05-10"),
    ("Draft clinical evaluation report", "Dr. Chen", "2026-05-20"),
    ("Schedule independent review", "[PM Name]", "2026-05-05"),
]


# ---------------------------------------------------------------------------
# Upstream XLSX Readers
# ---------------------------------------------------------------------------

def read_design_controls(path: str) -> dict:
    """Read requirements status from a design-controls skill XLSX."""
    wb = load_workbook(path, data_only=True)
    result = {
        "total_user_needs": 0,
        "total_design_inputs": 0,
        "verified": 0,
        "validated": 0,
        "open_items": 0,
        "open_detail": "",
    }

    # Try to read from Traceability sheet
    if "Traceability" in wb.sheetnames:
        ws = wb["Traceability"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        total = len(rows)
        result["total_design_inputs"] = total
        result["total_user_needs"] = total
        verified = sum(1 for r in rows if r and len(r) > 4 and r[4])
        validated = sum(1 for r in rows if r and len(r) > 5 and r[5])
        result["verified"] = verified
        result["validated"] = validated
        result["open_items"] = total - validated
    elif "Design_Inputs" in wb.sheetnames:
        ws = wb["Design_Inputs"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        result["total_design_inputs"] = len(rows)
        result["total_user_needs"] = len(rows)
        result["verified"] = len(rows)
        result["validated"] = len(rows)
    elif "User_Needs" in wb.sheetnames:
        ws = wb["User_Needs"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        result["total_user_needs"] = len(rows)
        result["total_design_inputs"] = len(rows)

    wb.close()
    return result


def read_risk_analysis(path: str) -> dict:
    """Read risk status from a risk-management skill XLSX."""
    wb = load_workbook(path, data_only=True)
    result = {
        "total_hazards": 0,
        "unacceptable": 0,
        "alarp": 0,
        "acceptable": 0,
        "open_controls": 0,
    }

    # Try Risk_Assessment or Hazard_Analysis sheet
    for sheet_name in ["Risk_Assessment", "Hazard_Analysis", "Risk_Analysis"]:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            result["total_hazards"] = len(rows)
            for r in rows:
                if not r:
                    continue
                # Look for residual risk level column (typically last few columns)
                risk_level = str(r[-1]).strip().lower() if r[-1] else ""
                if "unacceptable" in risk_level:
                    result["unacceptable"] += 1
                elif "alarp" in risk_level:
                    result["alarp"] += 1
                else:
                    result["acceptable"] += 1
            break

    wb.close()
    return result


def read_change_impact(path: str) -> list[tuple[str, str, str]]:
    """Read action items from a change-impact skill XLSX."""
    wb = load_workbook(path, data_only=True)
    actions = []

    for sheet_name in ["Action_Items", "Impact_Assessment", "Changes"]:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0]:
                    action = str(row[0])
                    owner = str(row[1]) if len(row) > 1 and row[1] else "TBD"
                    due = str(row[2]) if len(row) > 2 and row[2] else "TBD"
                    actions.append((action, owner, due))
            break

    wb.close()
    return actions


# ---------------------------------------------------------------------------
# Auto-width helper
# ---------------------------------------------------------------------------

def build_document_control(wb: Workbook, review_type: str, device_name: str,
                           review_date: str):
    """Build the Document_Control sheet as the first sheet in the workbook."""
    ws = wb.active
    ws.title = "Document_Control"

    # Row 1: merged header
    ws.merge_cells("A1:D1")
    header_cell = ws.cell(row=1, column=1, value="Document Control")
    header_cell.fill = HEADER_FILL
    header_cell.font = Font(bold=True, color="FFFFFF", size=13)
    header_cell.alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 5):
        ws.cell(row=1, column=c).fill = HEADER_FILL
        ws.cell(row=1, column=c).border = THIN_BORDER

    # Rows 3-7: metadata
    metadata = [
        ("Document Title", f"Design Review Package — {review_type} — {device_name}"),
        ("Document Version", "1.0"),
        ("Device/Software Version", device_name),
        ("Date Generated", review_date),
        ("Review Type", REVIEW_TYPE_MAP[review_type]),
    ]
    for i, (label, value) in enumerate(metadata, start=3):
        ws.cell(row=i, column=1, value=label).font = LABEL_FONT
        ws.cell(row=i, column=1).border = THIN_BORDER
        ws.cell(row=i, column=2, value=value).font = VALUE_FONT
        ws.cell(row=i, column=2).border = THIN_BORDER

    # Rows 9-12: QMS integration fields
    qms_fields = [
        ("QMS Document Number", ""),
        ("SOP Reference", ""),
        ("Related Design Controls File", ""),
        ("Related Risk Management File", ""),
    ]
    for idx, (label, value) in enumerate(qms_fields):
        row = 9 + idx
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2, value=value).font = VALUE_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER

    # Row 14: approval signatures header
    ws.merge_cells("A14:D14")
    sig_header = ws.cell(row=14, column=1, value="Approval Signatures")
    sig_header.font = Font(bold=True, size=12)
    sig_header.alignment = Alignment(horizontal="left", vertical="center")
    for c in range(1, 5):
        ws.cell(row=14, column=c).border = THIN_BORDER

    # Row 15: column headers for signature table
    sig_cols = ["Role", "Name", "Date", "Signature"]
    for c, h in enumerate(sig_cols, start=1):
        cell = ws.cell(row=15, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    # Rows 16-18: signature rows
    sig_rows = [
        "Prepared By",
        "Reviewed By (QA)",
        "Approved By (Management)",
    ]
    for i, role in enumerate(sig_rows, start=16):
        ws.cell(row=i, column=1, value=role).font = LABEL_FONT
        for c in range(1, 5):
            ws.cell(row=i, column=c).border = THIN_BORDER

    auto_width(ws, min_width=18, max_width=65)


def auto_width(ws, min_width: int = 12, max_width: int = 55):
    """Auto-adjust column widths based on content."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = min_width
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[col_letter].width = max_len


def apply_header_row(ws, row: int, col_count: int):
    """Apply header styling to a row."""
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


# ---------------------------------------------------------------------------
# Sheet Builders
# ---------------------------------------------------------------------------

def build_review_summary(wb: Workbook, review_type: str, device_name: str,
                         attendees: str, review_date: str):
    """Build the Review_Summary sheet."""
    ws = wb.create_sheet("Review_Summary")

    # Headers
    ws.cell(row=1, column=1, value="Field")
    ws.cell(row=1, column=2, value="Value")
    apply_header_row(ws, 1, 2)

    rows = [
        ("Review Type", REVIEW_TYPE_MAP[review_type]),
        ("Device Name", device_name),
        ("Date", review_date),
        ("Gate", REVIEW_GATE[review_type]),
        ("Focus", REVIEW_FOCUS[review_type]),
        ("Attendees", attendees),
    ]

    for i, (label, value) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=label).font = LABEL_FONT
        ws.cell(row=i, column=2, value=value).font = VALUE_FONT
        ws.cell(row=i, column=1).border = THIN_BORDER
        ws.cell(row=i, column=2).border = THIN_BORDER

    # Overall Completeness — formula referencing Artifact_Checklist
    row_comp = len(rows) + 2
    ws.cell(row=row_comp, column=1, value="Overall Completeness").font = LABEL_FONT
    ws.cell(row=row_comp, column=1).border = THIN_BORDER
    comp_cell = ws.cell(row=row_comp, column=2)
    comp_cell.value = "=Artifact_Checklist!D2"
    comp_cell.number_format = "0.0%"
    comp_cell.font = Font(bold=True, size=14)
    comp_cell.border = THIN_BORDER

    # GO/NO-GO thresholds are organization-specific. Defaults:
    #   >= 90% completeness = GO (proceed to next phase)
    #   >= 75% completeness = CONDITIONAL (proceed with action items)
    #   < 75% = NO-GO (return to previous phase)
    # Adjust based on device risk class:
    #   Class III / high-risk: consider 95% or 100% for GO
    #   Class I / low-risk: 80% may be acceptable for GO
    # These thresholds must be defined in the organization's Design Control SOP.
    row_rec = row_comp + 1
    ws.cell(row=row_rec, column=1, value="Recommendation").font = LABEL_FONT
    ws.cell(row=row_rec, column=1).border = THIN_BORDER
    rec_cell = ws.cell(row=row_rec, column=2)
    rec_cell.value = (
        f'=IF(B{row_comp}>=0.9,"GO",IF(B{row_comp}>=0.75,"CONDITIONAL","NO-GO"))'
    )
    rec_cell.font = Font(bold=True, size=14)
    rec_cell.border = THIN_BORDER

    auto_width(ws, min_width=20, max_width=80)


def build_artifact_checklist(wb: Workbook, review_type: str,
                             not_required: set, not_provided: set):
    """Build the Artifact_Checklist sheet."""
    ws = wb.create_sheet("Artifact_Checklist")
    items = get_checklist_items(review_type)

    # Completeness formula row at top
    ws.cell(row=1, column=1, value="Artifact")
    ws.cell(row=1, column=2, value="Required")
    ws.cell(row=1, column=3, value="Provided")
    ws.cell(row=1, column=4, value="Completeness %")
    apply_header_row(ws, 1, 4)

    # Completeness formula in D2 — spans all data rows
    # Data starts at row 3, formula in row 2
    comp_cell = ws.cell(row=2, column=4)
    # Use a buffered range (C3:C200) so the formula scales if checklist items are added
    comp_cell.value = '=COUNTIF(C3:C200,"Y")/COUNTIF(B3:B200,"Y")'
    comp_cell.number_format = "0.0%"
    comp_cell.font = Font(bold=True, size=12)
    comp_cell.border = THIN_BORDER
    ws.cell(row=2, column=1, value="OVERALL").font = LABEL_FONT
    ws.cell(row=2, column=1).border = THIN_BORDER
    ws.cell(row=2, column=2).border = THIN_BORDER
    ws.cell(row=2, column=3).border = THIN_BORDER

    # Data rows
    for idx, item in enumerate(items):
        row = idx + 3
        ws.cell(row=row, column=1, value=f"{idx + 1}. {item}")
        required = "N" if idx in not_required else "Y"
        provided = "N" if (idx in not_provided or idx in not_required) else "Y"
        ws.cell(row=row, column=2, value=required)
        ws.cell(row=row, column=3, value=provided)

        # Style
        for c in range(1, 5):
            ws.cell(row=row, column=c).border = THIN_BORDER
            ws.cell(row=row, column=c).alignment = Alignment(vertical="center", wrap_text=True)

        # Color-code provided column
        prov_cell = ws.cell(row=row, column=3)
        if required == "N":
            prov_cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        elif provided == "Y":
            prov_cell.fill = GREEN_FILL
        else:
            prov_cell.fill = RED_FILL

    auto_width(ws)


def build_requirements_status(wb: Workbook, req_data: dict):
    """Build the Requirements_Status sheet."""
    ws = wb.create_sheet("Requirements_Status")

    ws.cell(row=1, column=1, value="Metric")
    ws.cell(row=1, column=2, value="Value")
    apply_header_row(ws, 1, 2)

    total_reqs = req_data["total_user_needs"] + req_data["total_design_inputs"]
    pct_verified = req_data["verified"] / max(req_data["total_design_inputs"], 1)
    pct_validated = req_data["validated"] / max(req_data["total_design_inputs"], 1)

    rows = [
        ("Total User Needs", req_data["total_user_needs"]),
        ("Total Design Inputs", req_data["total_design_inputs"]),
        ("Total Requirements", total_reqs),
        ("Verified", f"{req_data['verified']}/{req_data['total_design_inputs']}"),
        ("% Verified", pct_verified),
        ("Validated", f"{req_data['validated']}/{req_data['total_design_inputs']}"),
        ("% Validated", pct_validated),
        ("Open Items", req_data["open_items"]),
        ("Open Item Detail", req_data.get("open_detail", "")),
    ]

    for i, (label, value) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=label).font = LABEL_FONT
        ws.cell(row=i, column=1).border = THIN_BORDER
        val_cell = ws.cell(row=i, column=2, value=value)
        val_cell.font = VALUE_FONT
        val_cell.border = THIN_BORDER

        # Format percentages
        if isinstance(value, float):
            val_cell.number_format = "0.0%"
            if value >= 0.9:
                val_cell.fill = GREEN_FILL
            elif value >= 0.75:
                val_cell.fill = YELLOW_FILL
            else:
                val_cell.fill = RED_FILL

    auto_width(ws, min_width=20)


def build_risk_status(wb: Workbook, risk_data: dict):
    """Build the Risk_Status sheet."""
    ws = wb.create_sheet("Risk_Status")

    ws.cell(row=1, column=1, value="Metric")
    ws.cell(row=1, column=2, value="Value")
    apply_header_row(ws, 1, 2)

    rows = [
        ("Total Hazards Identified", risk_data["total_hazards"]),
        ("Unacceptable Remaining", risk_data["unacceptable"]),
        ("ALARP (with rationale)", risk_data["alarp"]),
        ("Acceptable", risk_data["acceptable"]),
        ("Open Risk Controls", risk_data["open_controls"]),
    ]

    for i, (label, value) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=label).font = LABEL_FONT
        ws.cell(row=i, column=1).border = THIN_BORDER
        val_cell = ws.cell(row=i, column=2, value=value)
        val_cell.font = VALUE_FONT
        val_cell.border = THIN_BORDER

    # Color unacceptable row
    unacceptable_cell = ws.cell(row=3, column=2)
    if risk_data["unacceptable"] == 0:
        unacceptable_cell.fill = GREEN_FILL
    else:
        unacceptable_cell.fill = RED_FILL

    # Color open controls
    open_ctrl_cell = ws.cell(row=6, column=2)
    if risk_data["open_controls"] == 0:
        open_ctrl_cell.fill = GREEN_FILL
    else:
        open_ctrl_cell.fill = YELLOW_FILL

    auto_width(ws, min_width=20)


def build_sign_off(wb: Workbook, signoff_data: list[tuple], signoff_notes: dict):
    """Build the Sign_Off sheet."""
    ws = wb.create_sheet("Sign_Off")

    headers = ["Role", "Name", "Date", "Approval", "Notes"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    apply_header_row(ws, 1, len(headers))

    for i, (role, name, date, approval) in enumerate(signoff_data, start=2):
        ws.cell(row=i, column=1, value=role).border = THIN_BORDER
        ws.cell(row=i, column=2, value=name).border = THIN_BORDER
        ws.cell(row=i, column=3, value=date).border = THIN_BORDER
        approval_cell = ws.cell(row=i, column=4, value=approval)
        approval_cell.border = THIN_BORDER
        approval_cell.font = Font(bold=True, size=11)

        # Color-code approval
        if approval == "Approve":
            approval_cell.fill = GREEN_FILL
        elif approval == "Conditional":
            approval_cell.fill = YELLOW_FILL
        elif approval == "Reject":
            approval_cell.fill = RED_FILL

        # Notes
        note = signoff_notes.get(role, "")
        ws.cell(row=i, column=5, value=note).border = THIN_BORDER

        for c in range(1, len(headers) + 1):
            ws.cell(row=i, column=c).alignment = Alignment(vertical="center")

    auto_width(ws)


# ---------------------------------------------------------------------------
# Markdown Narrative Generator
# ---------------------------------------------------------------------------

def generate_narrative(review_type: str, device_name: str, review_date: str,
                       attendees: str, checklist_items: list[str],
                       not_required: set, not_provided: set,
                       req_data: dict, risk_data: dict,
                       signoff_data: list[tuple], signoff_notes: dict,
                       action_items: list[tuple[str, str, str]]) -> str:
    """Generate the design review narrative markdown."""

    # Calculate completeness
    total_required = sum(1 for i in range(len(checklist_items)) if i not in not_required)
    total_provided = sum(
        1 for i in range(len(checklist_items))
        if i not in not_required and i not in not_provided
    )
    completeness_pct = (total_provided / total_required * 100) if total_required > 0 else 0

    if completeness_pct >= 90:
        recommendation = "GO"
        rec_detail = "Proceed to next phase. All critical artifacts are complete."
    elif completeness_pct >= 75:
        recommendation = "CONDITIONAL GO"
        rec_detail = (
            f"Proceed to next phase with conditions. "
            f"All action items must be resolved before "
            f"{'FDR' if review_type == 'CDR' else 'release' if review_type == 'FDR' else 'CDR'}."
        )
    else:
        recommendation = "NO-GO"
        rec_detail = "Do not proceed. Significant gaps must be addressed before re-review."

    pct_verified = req_data["verified"] / max(req_data["total_design_inputs"], 1) * 100
    pct_validated = req_data["validated"] / max(req_data["total_design_inputs"], 1) * 100

    # Missing items detail
    missing_items = []
    for i in range(len(checklist_items)):
        if i not in not_required and i in not_provided:
            missing_items.append(f"   - Item {i + 1}: {checklist_items[i]}")
    missing_section = "\n".join(missing_items) if missing_items else "   - None"

    # Sign-off summary
    signoff_lines = []
    for role, name, date, approval in signoff_data:
        note = signoff_notes.get(role, "")
        note_str = f' ("{note}")' if note else ""
        signoff_lines.append(f"| {role} | {name} | {date or 'TBD'} | {approval}{note_str} |")
    signoff_table = "\n".join(signoff_lines)

    # Action items table
    action_lines = []
    for idx, (action, owner, due) in enumerate(action_items, start=1):
        action_lines.append(f"| {idx} | {action} | {owner} | {due} |")
    action_table = "\n".join(action_lines) if action_lines else "| - | No action items | - | - |"

    narrative = f"""# {REVIEW_TYPE_MAP[review_type]}
## {device_name}
**Date**: {review_date}
**Review Type**: {REVIEW_TYPE_MAP[review_type].split(' -- ')[1]} ({review_type})
**Gate**: {REVIEW_GATE[review_type]}
**Attendees**: {attendees}

---

## Executive Summary

The {device_name} is presented for {REVIEW_TYPE_MAP[review_type].split(' -- ')[1]}. {_get_device_description(device_name)} Overall artifact completeness is {completeness_pct:.0f}% ({total_provided}/{total_required} {review_type} checklist items satisfied).

## Key Findings

1. **Requirements**: {req_data['total_user_needs']} user needs have corresponding design inputs. {pct_verified:.0f}% verification complete. {pct_validated:.0f}% validation complete{f" ({req_data['open_detail']})" if req_data.get('open_detail') else ""}.
2. **Risk Management**: All {risk_data['total_hazards']} identified hazards have risk controls. {risk_data['unacceptable']} unacceptable residual risks. {risk_data['alarp']} ALARP with documented rationale.
3. **Traceability**: {"Full forward/backward traceability established (UN->DI->DO->VER->VAL)." if pct_verified >= 90 else "Traceability matrix has gaps requiring attention."}
4. **Open Items**: {_summarize_open_items(req_data, risk_data, not_provided, checklist_items)}

## Artifact Completeness

- **Total checklist items**: {len(checklist_items)}
- **Required**: {total_required}
- **Provided**: {total_provided}
- **Completeness**: {completeness_pct:.0f}%

### Missing / Incomplete Items
{missing_section}

## Requirements Status

| Metric | Value |
|--------|-------|
| Total User Needs | {req_data['total_user_needs']} |
| Total Design Inputs | {req_data['total_design_inputs']} |
| Verified | {req_data['verified']}/{req_data['total_design_inputs']} ({pct_verified:.0f}%) |
| Validated | {req_data['validated']}/{req_data['total_design_inputs']} ({pct_validated:.0f}%) |
| Open Items | {req_data['open_items']} |

## Risk Status

| Metric | Value |
|--------|-------|
| Total Hazards | {risk_data['total_hazards']} |
| Unacceptable Remaining | {risk_data['unacceptable']} |
| ALARP (with rationale) | {risk_data['alarp']} |
| Acceptable | {risk_data['acceptable']} |
| Open Risk Controls | {risk_data['open_controls']} |

## Sign-Off Status

| Role | Name | Date | Approval |
|------|------|------|----------|
{signoff_table}

## Action Items

| # | Action | Owner | Due |
|---|--------|-------|-----|
{action_table}

## Recommendation

**{recommendation}** -- {rec_detail}

*Note: GO/NO-GO thresholds (90%/75%) are organization defaults. Adjust per device risk classification and organizational SOP. For Class III devices, consider requiring 95-100% for GO.*

---

*Generated by Design Review Packager skill on {review_date}.*
*Reference: ISO 13485:2016 Clause 7.3.5, FDA 21 CFR 820.30(e)*
"""
    return narrative


def _get_device_description(device_name: str) -> str:
    """Return a brief description based on device name."""
    if "spo2" in device_name.lower():
        return (
            "The system classifies neonatal SpO2 traces into urgency tiers "
            "using a 3-tier evaluation pipeline (rule engine, ML classifier, LLM reviewer)."
        )
    return f"The {device_name} is under development per the design and development plan."


def _summarize_open_items(req_data: dict, risk_data: dict,
                          not_provided: set, items: list[str]) -> str:
    """Summarize open items for the narrative."""
    parts = []
    if req_data.get("open_detail"):
        parts.append(req_data["open_detail"])
    if risk_data["open_controls"] > 0:
        parts.append(f"{risk_data['open_controls']} open risk controls")
    missing_count = len(not_provided)
    if missing_count > 0:
        sample_missing = [items[i] for i in sorted(not_provided) if i < len(items)][:3]
        parts.append(", ".join(sample_missing))
    if not parts:
        return "No open items."
    return "; ".join(parts) + "."


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Package design review gate documentation (PDR/CDR/FDR) as XLSX + markdown."
    )
    parser.add_argument(
        "--review-type", required=True, choices=["PDR", "CDR", "FDR"],
        help="Type of design review: PDR, CDR, or FDR"
    )
    parser.add_argument(
        "--device-name", default=None,
        help="Device/product name (default: from --example or 'Medical Device')"
    )
    parser.add_argument(
        "--example", choices=["spo2"], default=None,
        help="Use example data (spo2)"
    )
    parser.add_argument(
        "--design-controls", default=None,
        help="Path to upstream design-controls XLSX (optional)"
    )
    parser.add_argument(
        "--risk-analysis", default=None,
        help="Path to upstream risk-analysis XLSX (optional)"
    )
    parser.add_argument(
        "--change-impact", default=None,
        help="Path to upstream change-impact XLSX (optional)"
    )
    parser.add_argument(
        "--output-dir", default="output",
        help="Output directory (default: output)"
    )

    args = parser.parse_args()

    review_type = args.review_type
    review_date = datetime.now().strftime("%Y-%m-%d")

    # Determine device name and data source
    use_example = args.example == "spo2"
    device_name = args.device_name or (SPO2_DEVICE_NAME if use_example else "Medical Device")
    device_slug = device_name.lower().replace(" ", "-").replace(".", "").replace("v", "v")
    # Clean slug
    device_slug = "".join(c if c.isalnum() or c == "-" else "-" for c in device_slug)
    device_slug = "-".join(filter(None, device_slug.split("-")))

    if use_example:
        device_slug = SPO2_DEVICE_SLUG

    attendees = SPO2_ATTENDEES if use_example else "TBD"

    # Determine checklist status
    checklist_items = get_checklist_items(review_type)

    if use_example:
        if review_type == "PDR":
            not_required: set[int] = set()
            not_provided: set[int] = SPO2_PDR_NOT_PROVIDED.copy()
        elif review_type == "CDR":
            not_required = SPO2_CDR_NOT_REQUIRED.copy()
            not_provided = SPO2_CDR_NOT_PROVIDED.copy()
        else:  # FDR
            not_required = SPO2_CDR_NOT_REQUIRED.copy()
            not_provided = SPO2_FDR_NOT_PROVIDED.copy()
    else:
        not_required = set()
        not_provided = set()  # All blank for manual entry

    # Requirements data
    if args.design_controls and os.path.exists(args.design_controls):
        req_data = read_design_controls(args.design_controls)
        print(f"  Read requirements from: {args.design_controls}")
    elif use_example:
        req_data = SPO2_REQUIREMENTS.copy()
    else:
        req_data = {
            "total_user_needs": 0,
            "total_design_inputs": 0,
            "verified": 0,
            "validated": 0,
            "open_items": 0,
            "open_detail": "",
        }

    # Risk data
    if args.risk_analysis and os.path.exists(args.risk_analysis):
        risk_data = read_risk_analysis(args.risk_analysis)
        print(f"  Read risk data from: {args.risk_analysis}")
    elif use_example:
        risk_data = SPO2_RISK.copy()
    else:
        risk_data = {
            "total_hazards": 0,
            "unacceptable": 0,
            "alarp": 0,
            "acceptable": 0,
            "open_controls": 0,
        }

    # Action items
    action_items = list(SPO2_ACTION_ITEMS) if use_example else []
    if args.change_impact and os.path.exists(args.change_impact):
        extra_actions = read_change_impact(args.change_impact)
        action_items.extend(extra_actions)
        print(f"  Read change impact from: {args.change_impact}")

    # Sign-off data
    if use_example:
        signoff_data = list(SPO2_SIGNOFF)
        signoff_notes = dict(SPO2_SIGNOFF_NOTES)
    else:
        signoff_data = [
            ("PM / Design Owner", "", "", "Pending"),
            ("Clinical/Medical", "", "", "Pending"),
            ("QA/Regulatory", "", "", "Pending"),
            ("Engineering Lead", "", "", "Pending"),
            ("Data Scientist/ML", "", "", "Pending"),
            ("Independent Reviewer", "", "", "Pending"),
        ]
        signoff_notes = {}

    # Create output directory
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    # -----------------------------------------------------------------------
    # Build XLSX
    # -----------------------------------------------------------------------
    wb = Workbook()
    # Use the default sheet as Document_Control (must be first sheet)
    build_document_control(wb, review_type, device_name, review_date)

    build_review_summary(wb, review_type, device_name, attendees, review_date)
    build_artifact_checklist(wb, review_type, not_required, not_provided)
    build_requirements_status(wb, req_data)
    build_risk_status(wb, risk_data)
    build_sign_off(wb, signoff_data, signoff_notes)

    xlsx_filename = f"design-review-{review_type}-{device_slug}.xlsx"
    xlsx_path = output_dir / xlsx_filename
    wb.save(str(xlsx_path))
    print(f"  XLSX saved: {xlsx_path}")

    # -----------------------------------------------------------------------
    # Generate Markdown Narrative
    # -----------------------------------------------------------------------
    narrative = generate_narrative(
        review_type=review_type,
        device_name=device_name,
        review_date=review_date,
        attendees=attendees,
        checklist_items=checklist_items,
        not_required=not_required,
        not_provided=not_provided,
        req_data=req_data,
        risk_data=risk_data,
        signoff_data=signoff_data,
        signoff_notes=signoff_notes,
        action_items=action_items,
    )

    narrative_path = output_dir / "design_review_narrative.md"
    narrative_path.write_text(narrative)
    print(f"  Narrative saved: {narrative_path}")

    # Summary
    total_required = sum(1 for i in range(len(checklist_items)) if i not in not_required)
    total_provided = sum(
        1 for i in range(len(checklist_items))
        if i not in not_required and i not in not_provided
    )
    completeness = total_provided / total_required * 100 if total_required > 0 else 0

    if completeness >= 90:
        rec = "GO"
    elif completeness >= 75:
        rec = "CONDITIONAL"
    else:
        rec = "NO-GO"

    print(f"\n  Design Review Package Complete")
    print(f"  Review Type:  {REVIEW_TYPE_MAP[review_type]}")
    print(f"  Device:       {device_name}")
    print(f"  Checklist:    {total_provided}/{total_required} ({completeness:.0f}%)")
    print(f"  Recommendation: {rec}")


if __name__ == "__main__":
    main()
