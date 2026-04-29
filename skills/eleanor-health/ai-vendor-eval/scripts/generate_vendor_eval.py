#!/usr/bin/env python3
"""Generate AI vendor evaluation workbooks as XLSX.

Standalone script — only requires openpyxl.

Usage:
    python scripts/generate_vendor_eval.py --example eleanor
    python scripts/generate_vendor_eval.py --vendor "Vendor Name" --category "AI Voice Agent"
"""

import argparse
import os
import re
from datetime import date
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("Error: openpyxl is required. Install it with: pip install openpyxl")
    raise SystemExit(1)


# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGN = Alignment(vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Conditional formatting fills
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
GREEN_FONT = Font(color="006100")
RED_FONT = Font(color="9C0006")
YELLOW_FONT = Font(color="9C6500")


# ---------------------------------------------------------------------------
# Eleanor Health example data
# ---------------------------------------------------------------------------
def get_eleanor_example() -> dict[str, Any]:
    """Return hardcoded Eleanor Health AI voice agent vendor evaluation data."""
    vendor_name = "AI Voice Agent Vendor"
    category = "AI Voice Agent — Behavioral Health"

    evaluation_summary = [
        {
            "vendor_name": "Vendor A — VoiceCare AI",
            "category": category,
            "overall_score": "",  # Formula will be placed here
            "recommendation": "Conditional",
            "decision_date": date.today().strftime("%Y-%m-%d"),
            "owner": "Product Lead",
        },
    ]

    criteria_scoring = [
        # Clinical Safety
        {
            "criterion": "Crisis detection accuracy (suicidal ideation, self-harm)",
            "category": "Clinical Safety",
            "weight": 15,
            "score": 4,
            "notes": "Vendor claims 94% sensitivity on internal benchmark; independent validation pending",
        },
        {
            "criterion": "Clinical scope guardrails (no diagnosis, no prescribing)",
            "category": "Clinical Safety",
            "weight": 10,
            "score": 3,
            "notes": "Guardrails present but triggered inconsistently on edge-case prompts during demo",
        },
        {
            "criterion": "De-escalation protocol adherence",
            "category": "Clinical Safety",
            "weight": 10,
            "score": 4,
            "notes": "Follows motivational interviewing framework; warm handoff to crisis line confirmed",
        },
        # HIPAA & Compliance
        {
            "criterion": "BAA execution and terms",
            "category": "HIPAA & Compliance",
            "weight": 8,
            "score": 5,
            "notes": "Standard BAA signed; includes breach notification within 24 hours",
        },
        {
            "criterion": "42 CFR Part 2 compliance (SUD data)",
            "category": "HIPAA & Compliance",
            "weight": 8,
            "score": 3,
            "notes": "Vendor unfamiliar with Part 2 requirements; needs custom consent flow for SUD data",
        },
        {
            "criterion": "SOC 2 Type II certification",
            "category": "HIPAA & Compliance",
            "weight": 5,
            "score": 4,
            "notes": "SOC 2 Type II report dated 2025-09; annual audit cycle confirmed",
        },
        # Integration
        {
            "criterion": "athenahealth EHR integration (FHIR R4)",
            "category": "Integration",
            "weight": 8,
            "score": 3,
            "notes": "FHIR R4 API available; athenahealth-specific connector in beta, not GA",
        },
        {
            "criterion": "Salesforce Health Cloud integration",
            "category": "Integration",
            "weight": 5,
            "score": 4,
            "notes": "Native Salesforce app available; supports custom object mapping",
        },
        {
            "criterion": "Telephony platform compatibility (Twilio / RingCentral)",
            "category": "Integration",
            "weight": 5,
            "score": 5,
            "notes": "Twilio SIP trunk verified; RingCentral integration via partner",
        },
        # Cost & Value
        {
            "criterion": "Per-call cost vs. current cost-per-touchpoint",
            "category": "Cost & Value",
            "weight": 5,
            "score": 3,
            "notes": "$1.20/call vs. $4.80 current staff cost; volume discounts above 10K calls/month",
        },
        {
            "criterion": "Implementation cost and timeline",
            "category": "Cost & Value",
            "weight": 3,
            "score": 3,
            "notes": "$85K implementation; 12-week timeline including EHR integration",
        },
        {
            "criterion": "Total cost of ownership (3-year projection)",
            "category": "Cost & Value",
            "weight": 4,
            "score": 3,
            "notes": "3-year TCO ~$420K including licensing, integration, and support",
        },
        # Scalability
        {
            "criterion": "Multi-language support (Spanish, English minimum)",
            "category": "Scalability",
            "weight": 4,
            "score": 4,
            "notes": "English and Spanish GA; Portuguese and Haitian Creole on roadmap Q3 2026",
        },
        {
            "criterion": "Concurrent call capacity",
            "category": "Scalability",
            "weight": 3,
            "score": 5,
            "notes": "Tested to 500 concurrent calls; auto-scaling on AWS infrastructure",
        },
        # User Experience
        {
            "criterion": "Patient satisfaction scores from reference customers",
            "category": "User Experience",
            "weight": 4,
            "score": 4,
            "notes": "Reference customers report 4.2/5.0 patient satisfaction; NPS +32",
        },
        {
            "criterion": "Voice naturalness and latency",
            "category": "User Experience",
            "weight": 3,
            "score": 4,
            "notes": "Sub-500ms response latency; voice quality rated 'natural' by 78% of test users",
        },
    ]

    clinical_safety = [
        {
            "id": "CS-001",
            "requirement": "System must detect crisis language (suicidal ideation, self-harm, overdose intent) and immediately route to crisis line",
            "priority": "Must",
            "met": "Partial",
            "evidence": "Demo showed detection of explicit statements; missed indirect/coded language in 2 of 8 test cases",
            "gap_notes": "Require additional testing with behavioral health-specific crisis lexicon",
        },
        {
            "id": "CS-002",
            "requirement": "System must not provide clinical diagnoses, medication recommendations, or treatment plans",
            "priority": "Must",
            "met": "Yes",
            "evidence": "Tested 25 prompts attempting to elicit clinical advice; all correctly deflected",
            "gap_notes": "",
        },
        {
            "id": "CS-003",
            "requirement": "System must implement warm handoff to live clinician within 30 seconds when requested",
            "priority": "Must",
            "met": "Yes",
            "evidence": "Average handoff time 18 seconds in demo; queue management configurable",
            "gap_notes": "",
        },
        {
            "id": "CS-004",
            "requirement": "System must de-escalate distressed callers using trauma-informed language",
            "priority": "Must",
            "met": "Partial",
            "evidence": "Motivational interviewing framework present; trauma-informed care vocabulary needs customization",
            "gap_notes": "Need to train on Eleanor Health's clinical communication guidelines",
        },
        {
            "id": "CS-005",
            "requirement": "System must log all interactions for clinical quality review",
            "priority": "Must",
            "met": "Yes",
            "evidence": "Full transcript and audio recording with configurable retention; HIPAA-compliant storage",
            "gap_notes": "",
        },
        {
            "id": "CS-006",
            "requirement": "System must not continue conversation if patient reports active medical emergency",
            "priority": "Must",
            "met": "Yes",
            "evidence": "Emergency detection triggers 911 routing script; tested with 5 emergency scenarios",
            "gap_notes": "",
        },
        {
            "id": "CS-007",
            "requirement": "System must respect patient opt-out and Do Not Call preferences",
            "priority": "Should",
            "met": "Yes",
            "evidence": "Opt-out recorded in real time; syncs to DNC list within 1 hour",
            "gap_notes": "",
        },
        {
            "id": "CS-008",
            "requirement": "System must identify and appropriately handle minors (under 18)",
            "priority": "Should",
            "met": "No",
            "evidence": "No age verification or minor-specific workflow exists",
            "gap_notes": "Need custom workflow for minor callers; consider parental consent requirements",
        },
        {
            "id": "CS-009",
            "requirement": "System must detect substance use crisis indicators (overdose symptoms, withdrawal distress)",
            "priority": "Must",
            "met": "Partial",
            "evidence": "General crisis detection present; SUD-specific indicators not in training data",
            "gap_notes": "Vendor to add SUD crisis lexicon; validate with Eleanor clinical team",
        },
        {
            "id": "CS-010",
            "requirement": "System must not store or repeat back sensitive disclosures (assault, abuse) beyond the session",
            "priority": "Must",
            "met": "Yes",
            "evidence": "Session memory is ephemeral; sensitive topics flagged but not persisted in patient profile",
            "gap_notes": "",
        },
    ]

    hipaa_compliance = [
        {
            "id": "HC-001",
            "requirement": "Execute Business Associate Agreement (BAA) covering all PHI handling",
            "status": "Compliant",
            "evidence": "BAA signed 2026-03-15; covers voice data, transcripts, and metadata",
            "risk_if_unmet": "HIPAA violation; potential OCR enforcement action and fines up to $1.5M per category",
        },
        {
            "id": "HC-002",
            "requirement": "Encrypt PHI at rest (AES-256) and in transit (TLS 1.2+)",
            "status": "Compliant",
            "evidence": "AES-256 at rest confirmed; TLS 1.3 in transit; encryption keys managed via AWS KMS",
            "risk_if_unmet": "Data breach exposure; HIPAA Security Rule violation",
        },
        {
            "id": "HC-003",
            "requirement": "Maintain audit logs of all PHI access for minimum 6 years",
            "status": "Compliant",
            "evidence": "CloudTrail + application-level audit logs; 7-year retention configured",
            "risk_if_unmet": "Cannot demonstrate compliance during audit; HIPAA Administrative Safeguard violation",
        },
        {
            "id": "HC-004",
            "requirement": "42 CFR Part 2 compliant consent management for SUD treatment data",
            "status": "Non-Compliant",
            "evidence": "Vendor does not currently support Part 2 consent workflows",
            "risk_if_unmet": "Federal violation for unauthorized SUD data disclosure; criminal penalties possible",
        },
        {
            "id": "HC-005",
            "requirement": "Breach notification without unreasonable delay, no later than 60 days per HIPAA (45 CFR 164.404(b)); per BAA terms for vendor-to-CE notification",
            "status": "Compliant",
            "evidence": "BAA specifies 24-hour vendor notification to covered entity; incident response plan reviewed",
            "risk_if_unmet": "Delayed breach response; regulatory penalties and patient harm",
        },
        {
            "id": "HC-006",
            "requirement": "Role-based access control (RBAC) for all PHI systems",
            "status": "Compliant",
            "evidence": "RBAC implemented with quarterly access reviews; MFA required for all admin access",
            "risk_if_unmet": "Unauthorized PHI access; minimum necessary standard violation",
        },
        {
            "id": "HC-007",
            "requirement": "Data residency within United States (no offshore processing)",
            "status": "Compliant",
            "evidence": "All processing in AWS us-east-1 and us-west-2; no cross-border data transfer",
            "risk_if_unmet": "Potential state privacy law violations; patient trust concerns",
        },
        {
            "id": "HC-008",
            "requirement": "Patient right to access and delete voice recordings",
            "status": "Partial",
            "evidence": "Access request workflow exists; deletion workflow in development (ETA Q2 2026)",
            "risk_if_unmet": "HIPAA Right of Access violation; OCR enforcement priority area",
        },
        {
            "id": "HC-009",
            "requirement": "Annual HIPAA security risk assessment",
            "status": "Compliant",
            "evidence": "Most recent risk assessment dated 2025-11; third-party assessor (Coalfire)",
            "risk_if_unmet": "Non-compliance with HIPAA Security Rule; audit finding",
        },
        {
            "id": "HC-010",
            "requirement": "Workforce training on PHI handling and security awareness",
            "status": "Compliant",
            "evidence": "Annual training with 98% completion rate; phishing simulation program active",
            "risk_if_unmet": "Human error leading to breach; HIPAA Administrative Safeguard gap",
        },
        {
            "id": "HC-011",
            "requirement": "De-identification of data used for model training (Safe Harbor or Expert Determination)",
            "status": "Partial",
            "evidence": "Safe Harbor method applied to training data; re-identification risk assessment not completed",
            "risk_if_unmet": "Using PHI for model training without authorization; HIPAA Privacy Rule violation",
        },
        {
            "id": "HC-012",
            "requirement": "Subcontractor/sub-processor BAAs for all downstream vendors",
            "status": "Compliant",
            "evidence": "Sub-processor list provided; BAAs in place for AWS, Twilio, and transcription service",
            "risk_if_unmet": "Downstream breach without contractual protection; liability exposure",
        },
    ]

    integration_assessment = [
        {
            "system": "athenahealth EHR",
            "integration_type": "FHIR R4 API (Patient, Encounter, DocumentReference)",
            "feasibility": "Medium",
            "effort_estimate": "6-8 weeks",
            "dependencies": "athenahealth Marketplace approval; FHIR R4 sandbox access",
            "notes": "athenahealth connector in beta; may need custom adapter for Eleanor's workflow",
        },
        {
            "system": "Salesforce Health Cloud",
            "integration_type": "REST API + Custom Objects (Care Plan, Task)",
            "feasibility": "High",
            "effort_estimate": "3-4 weeks",
            "dependencies": "Salesforce Connected App setup; Eleanor custom object schema",
            "notes": "Native Salesforce app available; bi-directional sync for care coordination tasks",
        },
        {
            "system": "Twilio Telephony",
            "integration_type": "SIP Trunk + Programmable Voice API",
            "feasibility": "High",
            "effort_estimate": "1-2 weeks",
            "dependencies": "Twilio account configuration; SIP trunk provisioning",
            "notes": "Vendor has production Twilio integration; supports call recording and real-time transcription",
        },
        {
            "system": "Data Warehouse (Snowflake)",
            "integration_type": "Batch ETL via S3 + Snowpipe",
            "feasibility": "High",
            "effort_estimate": "2-3 weeks",
            "dependencies": "Snowflake staging schema; IAM role for S3 access",
            "notes": "Vendor exports call metadata and outcomes in Parquet format; daily batch load",
        },
        {
            "system": "Ops Platform (internal scheduling)",
            "integration_type": "REST API webhook for appointment reminders",
            "feasibility": "Medium",
            "effort_estimate": "4-5 weeks",
            "dependencies": "Internal API documentation; webhook endpoint security review",
            "notes": "Custom integration required; vendor has webhook framework but needs Eleanor-specific payload mapping",
        },
    ]

    pilot_design = [
        {
            "metric": "Appointment show rate",
            "baseline": "62%",
            "target": "72%",
            "measurement_method": "athenahealth appointment status report (kept vs. no-show)",
            "duration": "90 days",
            "success_threshold": ">= 70% show rate in pilot cohort",
            "owner": "Director of Operations",
        },
        {
            "metric": "Patient engagement rate (calls answered/completed)",
            "baseline": "35% (current outbound answer rate)",
            "target": "55%",
            "measurement_method": "Twilio call disposition logs + vendor dashboard",
            "duration": "90 days",
            "success_threshold": ">= 50% engagement rate",
            "owner": "Product Lead",
        },
        {
            "metric": "Crisis detection accuracy",
            "baseline": "N/A (new capability)",
            "target": "95% sensitivity",
            "measurement_method": "Clinical review of flagged vs. unflagged calls (random sample n=200)",
            "duration": "90 days",
            "success_threshold": ">= 90% sensitivity; zero missed critical events",
            "owner": "Chief Clinical Officer",
        },
        {
            "metric": "Patient satisfaction (post-call survey)",
            "baseline": "3.8/5.0 (current IVR satisfaction)",
            "target": "4.3/5.0",
            "measurement_method": "Post-call SMS survey (1-5 scale); minimum 30% response rate",
            "duration": "90 days",
            "success_threshold": ">= 4.0/5.0 with >= 30% response rate",
            "owner": "Patient Experience Manager",
        },
        {
            "metric": "Cost per successful patient touchpoint",
            "baseline": "$4.80/touchpoint (staff outreach)",
            "target": "$2.00/touchpoint",
            "measurement_method": "Total vendor cost / completed calls resulting in action (appointment, refill, etc.)",
            "duration": "90 days",
            "success_threshold": "<= $2.50/touchpoint",
            "owner": "Finance Lead",
        },
        {
            "metric": "Clinical escalation appropriateness",
            "baseline": "N/A (new capability)",
            "target": ">= 90% appropriate escalations",
            "measurement_method": "Clinical team review of all escalated calls; rate as appropriate/inappropriate",
            "duration": "90 days",
            "success_threshold": ">= 85% appropriate; zero missed escalations",
            "owner": "Clinical Quality Manager",
        },
        {
            "metric": "Staff time recaptured (hours/week)",
            "baseline": "0 hours (current state)",
            "target": "40 hours/week across pilot sites",
            "measurement_method": "Time study: staff hours on outbound calls pre/post pilot",
            "duration": "90 days",
            "success_threshold": ">= 30 hours/week recaptured",
            "owner": "Director of Operations",
        },
        {
            "metric": "System uptime and reliability",
            "baseline": "N/A",
            "target": "99.9% uptime",
            "measurement_method": "Vendor SLA dashboard + internal monitoring (PagerDuty)",
            "duration": "90 days",
            "success_threshold": ">= 99.5% uptime; zero unplanned outages > 30 min",
            "owner": "Engineering Lead",
        },
    ]

    contract_requirements = [
        {
            "term": "Business Associate Agreement",
            "requirement": "Fully executed BAA covering voice data, transcripts, metadata, and derived analytics",
            "vendor_response": "Standard BAA provided; covers all PHI categories",
            "acceptable": "Yes",
            "negotiation_notes": "",
        },
        {
            "term": "Data ownership and portability",
            "requirement": "Eleanor Health retains ownership of all patient data; vendor provides data export within 30 days of termination",
            "vendor_response": "Data ownership clause included; export in Parquet + JSON format",
            "acceptable": "Yes",
            "negotiation_notes": "",
        },
        {
            "term": "Model training data usage",
            "requirement": "Vendor may NOT use Eleanor patient data to train models for other customers without explicit written consent",
            "vendor_response": "Agreed to opt-out; default is opt-in for aggregate model improvement",
            "acceptable": "No",
            "negotiation_notes": "Require explicit opt-out as DEFAULT, not opt-in. Non-negotiable for Part 2 compliance.",
        },
        {
            "term": "SLA: uptime guarantee",
            "requirement": "99.9% uptime SLA with service credits for breach",
            "vendor_response": "99.5% standard SLA; 99.9% available on Enterprise tier at +15% cost",
            "acceptable": "No",
            "negotiation_notes": "Negotiate 99.9% into base contract; clinical use case requires higher reliability",
        },
        {
            "term": "Termination and wind-down",
            "requirement": "90-day termination for convenience with 180-day wind-down support",
            "vendor_response": "60-day termination; 90-day wind-down",
            "acceptable": "Yes",
            "negotiation_notes": "Acceptable; shorter than requested but operationally feasible",
        },
        {
            "term": "Liability cap",
            "requirement": "Liability cap of 2x annual contract value; unlimited for data breach and willful misconduct",
            "vendor_response": "1x annual contract value; unlimited for data breach only",
            "acceptable": "No",
            "negotiation_notes": "Push for 2x cap; add willful misconduct carve-out",
        },
        {
            "term": "Insurance requirements",
            "requirement": "Vendor carries $5M cyber liability insurance and $2M professional liability (E&O)",
            "vendor_response": "$5M cyber; $1M E&O",
            "acceptable": "No",
            "negotiation_notes": "Require $2M E&O minimum given clinical use case; verify certificate of insurance",
        },
        {
            "term": "Security incident response",
            "requirement": "Vendor notifies Eleanor within 24 hours of any security incident; provides root cause analysis within 5 business days",
            "vendor_response": "24-hour notification agreed; RCA within 10 business days",
            "acceptable": "Yes",
            "negotiation_notes": "10-day RCA acceptable; add requirement for interim status updates every 48 hours",
        },
        {
            "term": "Audit rights",
            "requirement": "Eleanor Health may audit vendor security controls annually; vendor provides SOC 2 Type II report",
            "vendor_response": "SOC 2 Type II provided annually; on-site audit with 30-day notice",
            "acceptable": "Yes",
            "negotiation_notes": "",
        },
        {
            "term": "Regulatory change management",
            "requirement": "Vendor adapts to regulatory changes (HIPAA, state privacy laws, 42 CFR Part 2) within 90 days of effective date",
            "vendor_response": "Best-effort compliance; no SLA on regulatory adaptation",
            "acceptable": "No",
            "negotiation_notes": "Require contractual SLA for regulatory compliance updates; critical for behavioral health",
        },
    ]

    return {
        "vendor_name": vendor_name,
        "category": category,
        "evaluation_summary": evaluation_summary,
        "criteria_scoring": criteria_scoring,
        "clinical_safety": clinical_safety,
        "hipaa_compliance": hipaa_compliance,
        "integration_assessment": integration_assessment,
        "pilot_design": pilot_design,
        "contract_requirements": contract_requirements,
    }


# ---------------------------------------------------------------------------
# Blank scaffold (for custom --vendor usage)
# ---------------------------------------------------------------------------
def get_blank_scaffold(vendor_name: str, category: str) -> dict[str, Any]:
    """Return a minimal scaffold with one placeholder row per sheet."""
    return {
        "vendor_name": vendor_name,
        "category": category,
        "evaluation_summary": [
            {
                "vendor_name": vendor_name,
                "category": category,
                "overall_score": "",
                "recommendation": "",
                "decision_date": date.today().strftime("%Y-%m-%d"),
                "owner": "<Owner>",
            },
        ],
        "criteria_scoring": [
            {
                "criterion": "<Enter evaluation criterion>",
                "category": "Clinical Safety",
                "weight": 10,
                "score": "",
                "notes": "<Evidence or notes>",
            },
        ],
        "clinical_safety": [
            {
                "id": "CS-001",
                "requirement": "<Enter clinical safety requirement>",
                "priority": "Must",
                "met": "",
                "evidence": "<Evidence>",
                "gap_notes": "<Gap description if not met>",
            },
        ],
        "hipaa_compliance": [
            {
                "id": "HC-001",
                "requirement": "<Enter HIPAA/compliance requirement>",
                "status": "",
                "evidence": "<Evidence>",
                "risk_if_unmet": "<Risk description>",
            },
        ],
        "integration_assessment": [
            {
                "system": "<System name>",
                "integration_type": "<Integration approach>",
                "feasibility": "",
                "effort_estimate": "<Estimate>",
                "dependencies": "<Dependencies>",
                "notes": "<Notes>",
            },
        ],
        "pilot_design": [
            {
                "metric": "<Metric name>",
                "baseline": "<Current value>",
                "target": "<Target value>",
                "measurement_method": "<How to measure>",
                "duration": "<Duration>",
                "success_threshold": "<Pass/fail threshold>",
                "owner": "<Owner>",
            },
        ],
        "contract_requirements": [
            {
                "term": "<Contract term>",
                "requirement": "<Your requirement>",
                "vendor_response": "<Vendor position>",
                "acceptable": "",
                "negotiation_notes": "<Notes>",
            },
        ],
    }


# ---------------------------------------------------------------------------
# Utility functions
# ---------------------------------------------------------------------------
def slugify(name: str) -> str:
    """Convert name to filename-safe slug."""
    slug = name.lower().strip()
    slug = re.sub(r"[^a-z0-9]+", "-", slug)
    slug = slug.strip("-")
    return slug


def style_header_row(ws, headers: list[str]) -> None:
    """Write and style the header row."""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def auto_width(ws, min_width: int = 12, max_width: int = 60) -> None:
    """Auto-adjust column widths based on content."""
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            if cell.value:
                # Handle multiline — use longest line
                lines = str(cell.value).split("\n")
                line_max = max(len(line) for line in lines)
                max_len = max(max_len, line_max)
        adjusted = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


def add_status_validation(ws, col_letter: str, row_start: int, row_end: int,
                          options: str = "Pass,Fail,Pending",
                          error_title: str = "Invalid Status",
                          error_msg: str = "Select a valid status.") -> None:
    """Add dropdown validation to a column."""
    dv = DataValidation(
        type="list",
        formula1=f'"{options}"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle=error_title,
        error=error_msg,
    )
    dv.sqref = f"{col_letter}{row_start}:{col_letter}{row_end}"
    ws.add_data_validation(dv)


def add_status_conditional_formatting(ws, col_letter: str, row_start: int, row_end: int,
                                      green_val: str = "Pass",
                                      red_val: str = "Fail",
                                      yellow_val: str = "Pending") -> None:
    """Add green/red/yellow conditional formatting to a column."""
    cell_range = f"{col_letter}{row_start}:{col_letter}{row_end}"
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="equal", formula=[f'"{green_val}"'], fill=GREEN_FILL, font=GREEN_FONT),
    )
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="equal", formula=[f'"{red_val}"'], fill=RED_FILL, font=RED_FONT),
    )
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="equal", formula=[f'"{yellow_val}"'], fill=YELLOW_FILL, font=YELLOW_FONT),
    )


def style_data_rows(ws, start_row: int, end_row: int, num_cols: int) -> None:
    """Apply alignment and borders to data rows."""
    for row in range(start_row, end_row + 1):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = CELL_ALIGN
            cell.border = THIN_BORDER


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------
def build_evaluation_summary(wb: Workbook, data: dict[str, Any]) -> None:
    """Build the Evaluation_Summary sheet as the first sheet (uses wb.active).

    High-level scorecard with recommendation conditional formatting and a
    COUNTIF-based overall score that averages weighted scores from
    the Criteria_Scoring sheet.
    """
    ws = wb.active
    ws.title = "Evaluation_Summary"

    headers = ["Vendor Name", "Category", "Overall Score", "Recommendation",
               "Decision Date", "Owner"]
    style_header_row(ws, headers)

    rows = data["evaluation_summary"]
    for i, row in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=row["vendor_name"])
        ws.cell(row=i, column=2, value=row["category"])
        # Overall Score: SUM of weighted scores / SUM of weights from Criteria_Scoring
        ws.cell(row=i, column=3,
                value='=IF(SUM(Criteria_Scoring!C:C)=0,"",SUM(Criteria_Scoring!E:E)/SUM(Criteria_Scoring!C:C)*100)')
        ws.cell(row=i, column=4, value=row["recommendation"])
        ws.cell(row=i, column=5, value=row["decision_date"])
        ws.cell(row=i, column=6, value=row["owner"])

    end_row = len(rows) + 1
    style_data_rows(ws, 2, end_row, len(headers))

    # Recommendation dropdown
    add_status_validation(ws, "D", 2, end_row,
                          options="Proceed,Conditional,Reject",
                          error_title="Invalid Recommendation",
                          error_msg="Recommendation must be Proceed, Conditional, or Reject.")

    # Conditional formatting on Recommendation column (D)
    add_status_conditional_formatting(ws, "D", 2, end_row,
                                      green_val="Proceed",
                                      red_val="Reject",
                                      yellow_val="Conditional")

    # Summary statistics below the table
    stats_row = end_row + 2
    ws.cell(row=stats_row, column=1, value="Scoring Breakdown")
    ws.cell(row=stats_row, column=1).font = Font(bold=True, size=12)

    ws.cell(row=stats_row + 1, column=1, value="Total Criteria:")
    ws.cell(row=stats_row + 1, column=1).font = Font(bold=True)
    ws.cell(row=stats_row + 1, column=2,
            value='=COUNTA(Criteria_Scoring!A:A)-1')

    ws.cell(row=stats_row + 2, column=1, value="Criteria Scored 4+:")
    ws.cell(row=stats_row + 2, column=1).font = Font(bold=True)
    ws.cell(row=stats_row + 2, column=2,
            value='=COUNTIF(Criteria_Scoring!D:D,">="&4)')

    ws.cell(row=stats_row + 3, column=1, value="Criteria Scored <=2:")
    ws.cell(row=stats_row + 3, column=1).font = Font(bold=True)
    ws.cell(row=stats_row + 3, column=2,
            value='=COUNTIF(Criteria_Scoring!D:D,"<="&2)')

    ws.cell(row=stats_row + 4, column=1, value="Clinical Safety Items Met:")
    ws.cell(row=stats_row + 4, column=1).font = Font(bold=True)
    ws.cell(row=stats_row + 4, column=2,
            value='=COUNTIF(Clinical_Safety!D:D,"Yes")')

    ws.cell(row=stats_row + 5, column=1, value="HIPAA Items Compliant:")
    ws.cell(row=stats_row + 5, column=1).font = Font(bold=True)
    ws.cell(row=stats_row + 5, column=2,
            value='=COUNTIF(HIPAA_Compliance!C:C,"Compliant")')

    ws.cell(row=stats_row + 6, column=1, value="Contract Terms Acceptable:")
    ws.cell(row=stats_row + 6, column=1).font = Font(bold=True)
    ws.cell(row=stats_row + 6, column=2,
            value='=COUNTIF(Contract_Requirements!D:D,"Yes")')

    auto_width(ws)


def build_criteria_scoring(wb: Workbook, data: list[dict]) -> None:
    """Build the Criteria_Scoring sheet — weighted evaluation matrix."""
    ws = wb.create_sheet("Criteria_Scoring")
    headers = ["Criterion", "Category", "Weight (%)", "Score (1-5)",
               "Weighted Score", "Evidence/Notes"]
    style_header_row(ws, headers)

    for i, row in enumerate(data, 2):
        ws.cell(row=i, column=1, value=row["criterion"])
        ws.cell(row=i, column=2, value=row["category"])
        ws.cell(row=i, column=3, value=row["weight"])
        if row["score"] != "":
            ws.cell(row=i, column=4, value=row["score"])
        # Weighted Score formula: =C{row}*D{row}/100
        ws.cell(row=i, column=5, value=f"=C{i}*D{i}/100")
        ws.cell(row=i, column=6, value=row["notes"])

    end_row = len(data) + 1
    style_data_rows(ws, 2, end_row, len(headers))

    # Category dropdown (column B)
    add_status_validation(
        ws, "B", 2, end_row,
        options="Clinical Safety,HIPAA & Compliance,Integration,Cost & Value,Scalability,User Experience",
        error_title="Invalid Category",
        error_msg="Select a valid evaluation category.",
    )

    # Score dropdown (column D): 1-5
    add_status_validation(
        ws, "D", 2, end_row,
        options="1,2,3,4,5",
        error_title="Invalid Score",
        error_msg="Score must be between 1 and 5.",
    )

    # Summary row below data
    summary_row = end_row + 1
    ws.cell(row=summary_row, column=2, value="TOTAL").font = Font(bold=True)
    ws.cell(row=summary_row, column=3, value=f"=SUM(C2:C{end_row})")
    ws.cell(row=summary_row, column=5, value=f"=SUM(E2:E{end_row})")
    style_data_rows(ws, summary_row, summary_row, len(headers))

    auto_width(ws)


def build_clinical_safety(wb: Workbook, data: list[dict]) -> None:
    """Build the Clinical_Safety sheet — safety requirements checklist."""
    ws = wb.create_sheet("Clinical_Safety")
    headers = ["Requirement ID", "Requirement", "Priority", "Met?",
               "Evidence", "Gap Notes"]
    style_header_row(ws, headers)

    for i, row in enumerate(data, 2):
        ws.cell(row=i, column=1, value=row["id"])
        ws.cell(row=i, column=2, value=row["requirement"])
        ws.cell(row=i, column=3, value=row["priority"])
        ws.cell(row=i, column=4, value=row["met"])
        ws.cell(row=i, column=5, value=row["evidence"])
        ws.cell(row=i, column=6, value=row["gap_notes"])

    end_row = len(data) + 1
    style_data_rows(ws, 2, end_row, len(headers))

    # Priority dropdown (column C)
    add_status_validation(
        ws, "C", 2, end_row,
        options="Must,Should,Nice",
        error_title="Invalid Priority",
        error_msg="Priority must be Must, Should, or Nice.",
    )

    # Met? dropdown (column D)
    add_status_validation(
        ws, "D", 2, end_row,
        options="Yes,No,Partial",
        error_title="Invalid Value",
        error_msg="Must be Yes, No, or Partial.",
    )

    # Conditional formatting on Met? column (D)
    add_status_conditional_formatting(ws, "D", 2, end_row,
                                      green_val="Yes",
                                      red_val="No",
                                      yellow_val="Partial")

    auto_width(ws)


def build_hipaa_compliance(wb: Workbook, data: list[dict]) -> None:
    """Build the HIPAA_Compliance sheet — compliance checklist."""
    ws = wb.create_sheet("HIPAA_Compliance")
    headers = ["Requirement ID", "Requirement", "Status", "Evidence",
               "Risk if Unmet"]
    style_header_row(ws, headers)

    for i, row in enumerate(data, 2):
        ws.cell(row=i, column=1, value=row["id"])
        ws.cell(row=i, column=2, value=row["requirement"])
        ws.cell(row=i, column=3, value=row["status"])
        ws.cell(row=i, column=4, value=row["evidence"])
        ws.cell(row=i, column=5, value=row["risk_if_unmet"])

    end_row = len(data) + 1
    style_data_rows(ws, 2, end_row, len(headers))

    # Status dropdown (column C)
    add_status_validation(
        ws, "C", 2, end_row,
        options="Compliant,Non-Compliant,Partial,N/A",
        error_title="Invalid Status",
        error_msg="Status must be Compliant, Non-Compliant, Partial, or N/A.",
    )

    # Conditional formatting on Status column (C)
    add_status_conditional_formatting(ws, "C", 2, end_row,
                                      green_val="Compliant",
                                      red_val="Non-Compliant",
                                      yellow_val="Partial")

    auto_width(ws)


def build_integration_assessment(wb: Workbook, data: list[dict]) -> None:
    """Build the Integration_Assessment sheet — tech stack compatibility."""
    ws = wb.create_sheet("Integration_Assessment")
    headers = ["System", "Integration Type", "Feasibility", "Effort Estimate",
               "Dependencies", "Notes"]
    style_header_row(ws, headers)

    for i, row in enumerate(data, 2):
        ws.cell(row=i, column=1, value=row["system"])
        ws.cell(row=i, column=2, value=row["integration_type"])
        ws.cell(row=i, column=3, value=row["feasibility"])
        ws.cell(row=i, column=4, value=row["effort_estimate"])
        ws.cell(row=i, column=5, value=row["dependencies"])
        ws.cell(row=i, column=6, value=row["notes"])

    end_row = len(data) + 1
    style_data_rows(ws, 2, end_row, len(headers))

    # Feasibility dropdown (column C)
    add_status_validation(
        ws, "C", 2, end_row,
        options="High,Medium,Low",
        error_title="Invalid Feasibility",
        error_msg="Feasibility must be High, Medium, or Low.",
    )

    # Conditional formatting on Feasibility column (C)
    add_status_conditional_formatting(ws, "C", 2, end_row,
                                      green_val="High",
                                      red_val="Low",
                                      yellow_val="Medium")

    auto_width(ws)


def build_pilot_design(wb: Workbook, data: list[dict]) -> None:
    """Build the Pilot_Design sheet — pilot plan template."""
    ws = wb.create_sheet("Pilot_Design")
    headers = ["Metric", "Baseline", "Target", "Measurement Method",
               "Duration", "Success Threshold", "Owner"]
    style_header_row(ws, headers)

    for i, row in enumerate(data, 2):
        ws.cell(row=i, column=1, value=row["metric"])
        ws.cell(row=i, column=2, value=row["baseline"])
        ws.cell(row=i, column=3, value=row["target"])
        ws.cell(row=i, column=4, value=row["measurement_method"])
        ws.cell(row=i, column=5, value=row["duration"])
        ws.cell(row=i, column=6, value=row["success_threshold"])
        ws.cell(row=i, column=7, value=row["owner"])

    end_row = len(data) + 1
    style_data_rows(ws, 2, end_row, len(headers))

    auto_width(ws)


def build_contract_requirements(wb: Workbook, data: list[dict]) -> None:
    """Build the Contract_Requirements sheet — contract terms checklist."""
    ws = wb.create_sheet("Contract_Requirements")
    headers = ["Term", "Requirement", "Vendor Response", "Acceptable?",
               "Negotiation Notes"]
    style_header_row(ws, headers)

    for i, row in enumerate(data, 2):
        ws.cell(row=i, column=1, value=row["term"])
        ws.cell(row=i, column=2, value=row["requirement"])
        ws.cell(row=i, column=3, value=row["vendor_response"])
        ws.cell(row=i, column=4, value=row["acceptable"])
        ws.cell(row=i, column=5, value=row["negotiation_notes"])

    end_row = len(data) + 1
    style_data_rows(ws, 2, end_row, len(headers))

    # Acceptable? dropdown (column D)
    add_status_validation(
        ws, "D", 2, end_row,
        options="Yes,No",
        error_title="Invalid Value",
        error_msg="Must be Yes or No.",
    )

    # Conditional formatting on Acceptable? column (D)
    cell_range = f"D2:D{end_row}"
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="equal", formula=['"Yes"'], fill=GREEN_FILL, font=GREEN_FONT),
    )
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="equal", formula=['"No"'], fill=RED_FILL, font=RED_FONT),
    )

    auto_width(ws)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def generate(data: dict[str, Any]) -> str:
    """Generate the XLSX workbook and return the output path."""
    wb = Workbook()

    # Build Evaluation_Summary first (uses wb.active, stays at tab index 0)
    build_evaluation_summary(wb, data)

    # Build all 6 remaining sheets
    build_criteria_scoring(wb, data["criteria_scoring"])
    build_clinical_safety(wb, data["clinical_safety"])
    build_hipaa_compliance(wb, data["hipaa_compliance"])
    build_integration_assessment(wb, data["integration_assessment"])
    build_pilot_design(wb, data["pilot_design"])
    build_contract_requirements(wb, data["contract_requirements"])

    # Set Evaluation_Summary as the active/first-visible sheet
    wb.active = 0

    # Save
    slug = slugify(data["vendor_name"])
    output_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "output")
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, f"vendor-eval-{slug}.xlsx")
    wb.save(output_path)
    return output_path


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Generate AI vendor evaluation workbook (XLSX)."
    )
    parser.add_argument(
        "--example",
        choices=["eleanor"],
        help="Generate a pre-populated example (eleanor = Eleanor Health AI Voice Agent).",
    )
    parser.add_argument(
        "--vendor",
        type=str,
        help="Vendor name for the evaluation.",
    )
    parser.add_argument(
        "--category",
        type=str,
        default="AI Vendor",
        help="Vendor category (default: AI Vendor).",
    )
    args = parser.parse_args()

    if args.example == "eleanor":
        data = get_eleanor_example()
    elif args.vendor:
        data = get_blank_scaffold(args.vendor, args.category)
    else:
        parser.error("Provide --example eleanor or --vendor 'Vendor Name'.")

    output_path = generate(data)
    print(f"Generated: {output_path}")
    sheets = [
        "Evaluation_Summary", "Criteria_Scoring", "Clinical_Safety",
        "HIPAA_Compliance", "Integration_Assessment", "Pilot_Design",
        "Contract_Requirements",
    ]
    print(f"  Sheets: {', '.join(sheets)}")
    print(f"  Vendor: {data['vendor_name']}")
    print(f"  Category: {data['category']}")
    record_counts = (
        f"{len(data['criteria_scoring'])} criteria, "
        f"{len(data['clinical_safety'])} safety reqs, "
        f"{len(data['hipaa_compliance'])} compliance reqs, "
        f"{len(data['integration_assessment'])} integrations, "
        f"{len(data['pilot_design'])} pilot metrics, "
        f"{len(data['contract_requirements'])} contract terms"
    )
    print(f"  Records: {record_counts}")


if __name__ == "__main__":
    main()
